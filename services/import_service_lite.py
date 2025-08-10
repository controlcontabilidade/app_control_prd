#!/usr/bin/env python3
"""
Serviço de Importação de Clientes Sem Pandas
Permite importar clientes de planilhas XLSX usando apenas openpyxl
"""

# Verificar dependências
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from datetime import datetime
import uuid
from typing import List, Dict, Tuple, Optional
import re

class ImportServiceLite:
    """Versão leve do ImportService sem pandas"""
    
    def __init__(self, storage_service):
        self.storage_service = storage_service
        self.openpyxl_available = OPENPYXL_AVAILABLE
        
        if not self.openpyxl_available:
            print("⚠️ ImportServiceLite inicializado sem openpyxl - funcionalidade limitada")
        else:
            print("✅ ImportServiceLite inicializado com openpyxl")
        
    def is_available(self) -> bool:
        """Verifica se o serviço de importação está disponível"""
        return self.openpyxl_available
        
    def read_excel_file(self, file_path: str) -> List[Dict]:
        """Lê arquivo Excel e retorna lista de dicionários"""
        if not self.openpyxl_available:
            raise ImportError("openpyxl não está disponível. Funcionalidade de importação desabilitada.")
            
        try:
            print(f"📊 Lendo arquivo Excel: {file_path}")
            
            # Carregar workbook
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet = workbook.active
            
            # Obter dados
            data = []
            headers = []
            
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_num == 1:
                    # Primeira linha são os cabeçalhos
                    headers = [str(cell).strip() if cell is not None else '' for cell in row]
                    continue
                
                # Criar dicionário com dados da linha
                row_data = {}
                for col_num, cell_value in enumerate(row):
                    if col_num < len(headers) and headers[col_num]:
                        value = str(cell_value).strip() if cell_value is not None else ''
                        row_data[headers[col_num]] = value
                
                # Pular linhas vazias
                if any(row_data.values()):
                    data.append(row_data)
            
            workbook.close()
            print(f"✅ Arquivo lido com sucesso: {len(data)} linhas encontradas")
            return data
            
        except Exception as e:
            print(f"❌ Erro ao ler arquivo Excel: {e}")
            raise e
    
    def normalize_column_names(self, data: List[Dict]) -> List[Dict]:
        """Normaliza nomes das colunas para corresponder ao sistema SIGEC"""
        column_mapping = {
            # Bloco 1: Informações da Pessoa Jurídica (novos nomes SIGEC)
            'NOME DA EMPRESA': 'nomeEmpresa',
            'RAZÃO SOCIAL NA RECEITA': 'razaoSocialReceita',
            'NOME FANTASIA NA RECEITA': 'nomeFantasiaReceita',
            'CNPJ': 'cnpj',
            'PERFIL': 'perfil',
            'INSCRIÇÃO ESTADUAL': 'inscEst',
            'INSC. EST.': 'inscEst',  # Compatibilidade
            'INSCRIÇÃO MUNICIPAL': 'inscMun',
            'INSC. MUN.': 'inscMun',  # Compatibilidade
            'ESTADO': 'estado',
            'CIDADE': 'cidade',
            'REGIME FEDERAL': 'regimeFederal',
            'TRIBUTAÇÃO': 'regimeFederal',  # Compatibilidade - mapear para regimeFederal
            'REGIME ESTADUAL': 'regimeEstadual',
            'SEGMENTO': 'segmento',
            'ATIVIDADE': 'atividade',
            
            # Bloco 2: Serviços Prestados pela Control
            'SERVIÇO CT': 'ct',
            'CT': 'ct',  # Compatibilidade
            'SERVIÇO FS': 'fs',
            'FS': 'fs',  # Compatibilidade
            'SERVIÇO DP': 'dp',
            'DP': 'dp',  # Compatibilidade
            'SERVIÇO BPO FINANCEIRO': 'bpoFinanceiro',
            'RESPONSÁVEL PELOS SERVIÇOS': 'responsavelServicos',
            'DONO / RESP.': 'responsavelServicos',  # Compatibilidade
            'DATA INÍCIO DOS SERVIÇOS': 'dataInicioServicos',
            'MÊS/ANO DE  INÍCIO': 'dataInicioServicos',  # Compatibilidade
            
            # Códigos dos Sistemas (Bloco 2)
            'CÓDIGO FORTES CT': 'codFortesCt',
            'COD. FORTES CT': 'codFortesCt',  # Compatibilidade
            'CÓDIGO FORTES FS': 'codFortesFs',
            'COD. FORTES FS': 'codFortesFs',  # Compatibilidade
            'CÓDIGO FORTES PS': 'codFortesPs',
            'COD. FORTES PS': 'codFortesPs',  # Compatibilidade
            'CÓDIGO DOMÍNIO': 'codDominio',
            'COD. DOMÍNIO': 'codDominio',  # Compatibilidade
            'SISTEMA UTILIZADO': 'sistemaUtilizado',
            'MÓDULO SPED TRIER': 'moduloSpedTrier',
            
            # Bloco 3: Quadro Societário (novos campos SIGEC)
            'SÓCIO 1 NOME': 'socio1_nome',
            'SÓCIO': 'socio1_nome',  # Compatibilidade
            'SÓCIO 1 CPF': 'socio1_cpf',
            'SÓCIO 1 DATA NASCIMENTO': 'socio1_nascimento',
            'SÓCIO 1 ADMINISTRADOR': 'socio1_admin',
            'SÓCIO 1 COTAS': 'socio1_cotas',
            'SÓCIO 1 RESPONSÁVEL LEGAL': 'socio1_resp_legal',
            
            # Compatibilidade com campos antigos de sócios
            'SÓCIO.1': 'socio2',
            'SÓCIO.2': 'socio3',
            'SÓCIO.3': 'socio4',
            
            # Bloco 4: Contatos
            'TELEFONE FIXO': 'telefoneFixo',
            'TELEFONE CELULAR': 'telefoneCelular',
            'WHATSAPP': 'whatsapp',
            'EMAIL PRINCIPAL': 'emailPrincipal',
            'EMAIL SECUNDÁRIO': 'emailSecundario',
            'RESPONSÁVEL IMEDIATO': 'responsavelImediato',
            'EMAILS DOS SÓCIOS': 'emailsSocios',
            'E-MAILS': 'emailsSocios',  # Compatibilidade
            'CONTATO CONTADOR': 'contatoContador',
            'TELEFONE CONTADOR': 'telefoneContador',
            'EMAIL CONTADOR': 'emailContador',
            
            # Bloco 5: Sistemas e Acessos
            'SISTEMA PRINCIPAL': 'sistemaPrincipal',
            'VERSÃO DO SISTEMA': 'versaoSistema',
            'CÓDIGO ACESSO SIMPLES NACIONAL': 'codAcessoSimples',
            'COD. ACESSO SIMPLES': 'codAcessoSimples',  # Compatibilidade
            'CPF/CNPJ PARA ACESSO': 'cpfCnpjAcesso',
            'CPF OU CNPJ': 'cpfCnpjAcesso',  # Compatibilidade
            'PORTAL CLIENTE ATIVO': 'portalClienteAtivo',
            'INTEGRAÇÃO DOMÍNIO': 'integracaoDominio',
            'SISTEMA ONVIO': 'sistemaOnvio',
            
            # Bloco 6: Senhas e Credenciais
            'ACESSO ISS': 'acessoIss',
            'SENHA ISS': 'senhaIss',
            'ACESSO SEFIN': 'acessoSefin',
            'SENHA SEFIN': 'senhaSefin',
            'ACESSO SEUMA': 'acessoSeuma',
            'SENHA SEUMA': 'senhaSeuma',
            'ACESSO EMPWEB': 'acessoEmpWeb',
            'ACESSO EMP. WEB': 'acessoEmpWeb',  # Compatibilidade
            'SENHA EMPWEB': 'senhaEmpWeb',
            'SENHA EMP. WEB': 'senhaEmpWeb',  # Compatibilidade
            'ACESSO FAP/INSS': 'acessoFapInss',
            'SENHA FAP/INSS': 'senhaFapInss',
            'ACESSO CRF': 'acessoCrf',
            'SENHA CRF': 'senhaCrf',
            'EMAIL GESTOR': 'emailGestor',
            'SENHA EMAIL GESTOR': 'senhaEmailGestor',
            'ANVISA GESTOR': 'anvisaGestor',
            'ANVISA EMPRESA': 'anvisaEmpresa',
            'ACESSO IBAMA': 'acessoIbama',
            'SENHA IBAMA': 'senhaIbama',
            'ACESSO SEMACE': 'acessoSemace',
            'SENHA SEMACE': 'senhaSemace',
            
            # Bloco 7: Procurações
            'PROCURAÇÃO RFB': 'procRfb',
            'DATA PROCURAÇÃO RFB': 'procRfbData',
            'PROCURAÇÃO RECEITA ESTADUAL': 'procRc',
            'PROC. RC': 'procRc',  # Compatibilidade
            'DATA PROCURAÇÃO RC': 'procRcData',
            'PROCURAÇÃO CAIXA ECONÔMICA': 'procCx',
            'PROC. CX': 'procCx',  # Compatibilidade
            'DATA PROCURAÇÃO CX': 'procCxData',
            'PROCURAÇÃO PREVIDÊNCIA SOCIAL': 'procSw',
            'PROC. SW': 'procSw',  # Compatibilidade
            'DATA PROCURAÇÃO SW': 'procSwData',
            'PROCURAÇÃO MUNICIPAL': 'procMunicipal',
            'DATA PROCURAÇÃO MUNICIPAL': 'procMunicipalData',
            'OUTRAS PROCURAÇÕES': 'outrasProc',
            'OBSERVAÇÕES PROCURAÇÕES': 'obsProcuracoes',
            
            # Bloco 8: Observações e Dados Adicionais (somente campos mantidos)
            'STATUS DO CLIENTE': 'statusCliente',
            'ÚLTIMA ATUALIZAÇÃO': 'ultimaAtualizacao',
            'RESPONSÁVEL ATUALIZAÇÃO': 'responsavelAtualizacao'
        }
        
        normalized_data = []
        for row in data:
            normalized_row = {}
            for original_col, value in row.items():
                mapped_col = column_mapping.get(original_col.upper(), original_col)
                normalized_row[mapped_col] = value
            normalized_data.append(normalized_row)
        
        return normalized_data
    
    def clean_data(self, data: List[Dict]) -> List[Dict]:
        """Limpa e padroniza os dados"""
        cleaned_data = []
        
        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                if value is None:
                    cleaned_row[key] = ''
                else:
                    # Converter para string e limpar
                    str_value = str(value).strip()
                    # Remover quebras de linha e espaços extras
                    str_value = re.sub(r'\s+', ' ', str_value)
                    cleaned_row[key] = str_value
            
            cleaned_data.append(cleaned_row)
        
        return cleaned_data
    
    def convert_boolean_fields(self, row: Dict) -> Dict:
        """Converte campos SIM/NÃO para boolean - SIGEC atualizado"""
        boolean_fields = [
            # Serviços
            'ct', 'fs', 'dp', 'bpoFinanceiro',
            # Sistemas e acessos
            'portalClienteAtivo', 'integracaoDominio', 'sistemaOnvio',
            # Sócios
            'socio1_admin', 'socio1_resp_legal',
            # Procurações
            'procRfb', 'procRc', 'procCx', 'procSw', 'procMunicipal',
            # Sistema
            'ativo'
        ]
        
        converted_row = row.copy()
        
        for field in boolean_fields:
            if field in converted_row:
                value = str(converted_row[field]).upper().strip()
                if value in ['SIM', 'S', 'TRUE', '1', 'VERDADEIRO', 'YES']:
                    converted_row[field] = True
                elif value in ['NÃO', 'NAO', 'N', 'FALSE', '0', 'FALSO', 'NO']:
                    converted_row[field] = False
                else:
                    converted_row[field] = False  # Default para False
        
        return converted_row
    
    def generate_unique_id(self) -> str:
        """Gera ID único para o cliente"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_suffix = str(uuid.uuid4())[:8]
        return f"IMP_{timestamp}_{unique_suffix}"
    
    def process_row_to_client(self, row: Dict, row_index: int) -> Optional[Dict]:
        """Processa uma linha do Excel para formato de cliente SIGEC"""
        try:
            # Verificar se tem nome da empresa (obrigatório)
            nome_empresa = row.get('nomeEmpresa', '').strip()
            if not nome_empresa:
                print(f"⚠️ Linha {row_index}: Nome da empresa é obrigatório")
                return None
            
            # Converter campos boolean
            row = self.convert_boolean_fields(row)
            
            # Criar dados do cliente SIGEC completo
            client_data = {
                'id': self.generate_unique_id(),
                'nomeEmpresa': nome_empresa,
                'criadoEm': datetime.now().isoformat(),
                
                # Bloco 1: Informações da Pessoa Jurídica
                'razaoSocialReceita': row.get('razaoSocialReceita', ''),
                'nomeFantasiaReceita': row.get('nomeFantasiaReceita', ''),
                'cnpj': row.get('cnpj', ''),
                'perfil': row.get('perfil', ''),
                'inscEst': row.get('inscEst', ''),
                'inscMun': row.get('inscMun', ''),
                'estado': row.get('estado', ''),
                'cidade': row.get('cidade', ''),
                'regimeFederal': row.get('regimeFederal', row.get('tributacao', '')),
                'regimeEstadual': row.get('regimeEstadual', ''),
                'segmento': row.get('segmento', ''),
                'atividade': row.get('atividade', ''),
                
                # Bloco 2: Serviços Prestados pela Control
                'ct': row.get('ct', False),
                'fs': row.get('fs', False),
                'dp': row.get('dp', False),
                'bpoFinanceiro': row.get('bpoFinanceiro', False),
                'responsavelServicos': row.get('responsavelServicos', row.get('donoResp', '')),
                'dataInicioServicos': row.get('dataInicioServicos', row.get('mesAnoInicio', '')),
                
                # Códigos dos Sistemas
                'codFortesCt': row.get('codFortesCt', ''),
                'codFortesFs': row.get('codFortesFs', ''),
                'codFortesPs': row.get('codFortesPs', ''),
                'codDominio': row.get('codDominio', ''),
                'sistemaUtilizado': row.get('sistemaUtilizado', ''),
                'moduloSpedTrier': row.get('moduloSpedTrier', ''),
                
                # Bloco 3: Quadro Societário
                'socio1_nome': row.get('socio1_nome', row.get('socio1', '')),
                'socio1_cpf': row.get('socio1_cpf', ''),
                'socio1_nascimento': row.get('socio1_nascimento', ''),
                'socio1_admin': row.get('socio1_admin', False),
                'socio1_cotas': row.get('socio1_cotas', ''),
                'socio1_resp_legal': row.get('socio1_resp_legal', False),
                
                # Campos de compatibilidade para sócios antigos
                'socio1': row.get('socio1_nome', row.get('socio1', '')),
                'socio2': row.get('socio2', ''),
                'socio3': row.get('socio3', ''),
                'socio4': row.get('socio4', ''),
                
                # Bloco 4: Contatos
                'telefoneFixo': row.get('telefoneFixo', ''),
                'telefoneCelular': row.get('telefoneCelular', ''),
                'whatsapp': row.get('whatsapp', ''),
                'emailPrincipal': row.get('emailPrincipal', ''),
                'emailSecundario': row.get('emailSecundario', ''),
                'responsavelImediato': row.get('responsavelImediato', ''),
                'emailsSocios': row.get('emailsSocios', row.get('emailsSocio', '')),
                'contatoContador': row.get('contatoContador', ''),
                'telefoneContador': row.get('telefoneContador', ''),
                'emailContador': row.get('emailContador', ''),
                
                # Bloco 5: Sistemas e Acessos
                'sistemaPrincipal': row.get('sistemaPrincipal', ''),
                'versaoSistema': row.get('versaoSistema', ''),
                'codAcessoSimples': row.get('codAcessoSimples', ''),
                'cpfCnpjAcesso': row.get('cpfCnpjAcesso', row.get('cpfOuCnpj', '')),
                'portalClienteAtivo': row.get('portalClienteAtivo', False),
                'integracaoDominio': row.get('integracaoDominio', False),
                'sistemaOnvio': row.get('sistemaOnvio', False),
                
                # Bloco 6: Senhas e Credenciais
                'acessoIss': row.get('acessoIss', ''),
                'senhaIss': row.get('senhaIss', ''),
                'acessoSefin': row.get('acessoSefin', ''),
                'senhaSefin': row.get('senhaSefin', ''),
                'acessoSeuma': row.get('acessoSeuma', ''),
                'senhaSeuma': row.get('senhaSeuma', ''),
                'acessoEmpWeb': row.get('acessoEmpWeb', ''),
                'senhaEmpWeb': row.get('senhaEmpWeb', ''),
                'acessoFapInss': row.get('acessoFapInss', ''),
                'senhaFapInss': row.get('senhaFapInss', ''),
                'acessoCrf': row.get('acessoCrf', ''),
                'senhaCrf': row.get('senhaCrf', ''),
                'emailGestor': row.get('emailGestor', ''),
                'senhaEmailGestor': row.get('senhaEmailGestor', ''),
                'anvisaGestor': row.get('anvisaGestor', ''),
                'anvisaEmpresa': row.get('anvisaEmpresa', ''),
                'acessoIbama': row.get('acessoIbama', ''),
                'senhaIbama': row.get('senhaIbama', ''),
                'acessoSemace': row.get('acessoSemace', ''),
                'senhaSemace': row.get('senhaSemace', ''),
                
                # Bloco 7: Procurações
                'procRfb': row.get('procRfb', False),
                'procRfbData': row.get('procRfbData', ''),
                'procRc': row.get('procRc', False),
                'procRcData': row.get('procRcData', ''),
                'procCx': row.get('procCx', False),
                'procCxData': row.get('procCxData', ''),
                'procSw': row.get('procSw', False),
                'procSwData': row.get('procSwData', ''),
                'procMunicipal': row.get('procMunicipal', False),
                'procMunicipalData': row.get('procMunicipalData', ''),
                'outrasProc': row.get('outrasProc', ''),
                'obsProcuracoes': row.get('obsProcuracoes', ''),
                
                # Bloco 8: Observações e Dados Adicionais (somente campos mantidos)
                'statusCliente': row.get('statusCliente', 'ATIVO'),
                'ultimaAtualizacao': row.get('ultimaAtualizacao', ''),
                'responsavelAtualizacao': row.get('responsavelAtualizacao', ''),
                
                # Campos internos do sistema
                'ativo': row.get('ativo', True),
                
                # Campos legados para compatibilidade
                'donoResp': row.get('responsavelServicos', row.get('donoResp', '')),
                'mesAnoInicio': row.get('dataInicioServicos', row.get('mesAnoInicio', '')),
                'emailsSocio': row.get('emailsSocios', row.get('emailsSocio', '')),
                'tributacao': row.get('regimeFederal', row.get('tributacao', '')),
                'cpfCnpj': row.get('cnpj', ''),
                'cpfOuCnpj': row.get('cpfCnpjAcesso', row.get('cpfOuCnpj', '')),
                'integradoDominio': row.get('integracaoDominio', False),
                'portalCliente': row.get('portalClienteAtivo', False),
                'onvio': row.get('sistemaOnvio', False)
            }
            
            return client_data
            
        except Exception as e:
            print(f"❌ Erro ao processar linha {row_index}: {e}")
            return None
    
    def validate_excel_structure(self, file_path: str) -> Tuple[bool, str]:
        """Valida a estrutura do arquivo Excel"""
        try:
            data = self.read_excel_file(file_path)
            
            if not data:
                return False, "Arquivo Excel está vazio ou não tem dados válidos"
            
            # Verificar se tem pelo menos a coluna obrigatória
            first_row = data[0]
            required_columns = ['NOME DA EMPRESA', 'nomeEmpresa']
            
            has_required = any(col in first_row for col in required_columns)
            
            if not has_required:
                return False, "Arquivo deve conter pelo menos a coluna 'NOME DA EMPRESA'"
            
            return True, "Estrutura válida"
            
        except Exception as e:
            return False, f"Erro ao validar arquivo: {str(e)}"
    
    def import_from_excel(self, file_path: str) -> Tuple[int, int, List[str]]:
        """Importa clientes de arquivo Excel"""
        sucessos = 0
        erros = 0
        lista_erros = []
        
        try:
            print("📊 Iniciando importação...")
            
            # Ler dados do Excel
            data = self.read_excel_file(file_path)
            print(f"📋 {len(data)} linhas encontradas")
            
            # Normalizar nomes das colunas
            data = self.normalize_column_names(data)
            
            # Limpar dados
            data = self.clean_data(data)
            
            # Processar cada linha
            for index, row in enumerate(data, start=2):  # Começar em 2 (linha 1 são cabeçalhos)
                try:
                    client_data = self.process_row_to_client(row, index)
                    
                    if client_data:
                        # Salvar cliente
                        if self.storage_service.save_client(client_data):
                            sucessos += 1
                            print(f"✅ Linha {index}: Cliente '{client_data['nomeEmpresa']}' importado")
                        else:
                            erros += 1
                            erro_msg = f"Linha {index}: Erro ao salvar cliente '{client_data.get('nomeEmpresa', 'N/A')}'"
                            lista_erros.append(erro_msg)
                            print(f"❌ {erro_msg}")
                    else:
                        erros += 1
                        erro_msg = f"Linha {index}: Dados inválidos ou incompletos"
                        lista_erros.append(erro_msg)
                        print(f"❌ {erro_msg}")
                        
                except Exception as e:
                    erros += 1
                    erro_msg = f"Linha {index}: Erro ao processar - {str(e)}"
                    lista_erros.append(erro_msg)
                    print(f"❌ {erro_msg}")
            
            print(f"📊 Importação concluída: {sucessos} sucessos, {erros} erros")
            return sucessos, erros, lista_erros
            
        except Exception as e:
            erro_msg = f"Erro geral na importação: {str(e)}"
            print(f"❌ {erro_msg}")
            return 0, 1, [erro_msg]

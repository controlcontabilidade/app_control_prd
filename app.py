# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
import json
import os
import gc  # Para otimização de memória
from datetime import datetime
from functools import wraps

# Importar otimizador de memória LITE para Render
try:
    from memory_optimizer_lite import RenderMemoryOptimizer, RENDER_MEMORY_SETTINGS, get_optimized_batch_size
    MEMORY_OPTIMIZER_AVAILABLE = True
    print("🚀 Render Memory Optimizer carregado")
    # Usar aliases para compatibilidade
    UltraMemoryOptimizer = RenderMemoryOptimizer
    ULTRA_MEMORY_SETTINGS = RENDER_MEMORY_SETTINGS
    get_ultra_optimized_batch_size = get_optimized_batch_size
except ImportError:
    try:
        from ultra_memory_optimizer import UltraMemoryOptimizer, ULTRA_MEMORY_SETTINGS, get_ultra_optimized_batch_size
        MEMORY_OPTIMIZER_AVAILABLE = True
        print("🧠 Ultra Memory Optimizer carregado")
    except ImportError:
        try:
            from memory_optimizer import MemoryOptimizer, MEMORY_OPTIMIZED_SETTINGS, get_optimized_batch_size
            MEMORY_OPTIMIZER_AVAILABLE = True
            print("🧠 Memory Optimizer carregado")
            # Usar versão padrão se ultra não estiver disponível
            UltraMemoryOptimizer = MemoryOptimizer
            ULTRA_MEMORY_SETTINGS = MEMORY_OPTIMIZED_SETTINGS
            get_ultra_optimized_batch_size = get_optimized_batch_size
        except ImportError:
            MEMORY_OPTIMIZER_AVAILABLE = False
            print("⚠️ Memory Optimizer não disponível")
            
            # Definir função fallback
            def get_ultra_optimized_batch_size():
                return 3  # Valor MUITO baixo para fallback

from services.google_sheets_service import GoogleSheetsService
from services.local_storage_service import LocalStorageService
from services.meeting_service import MeetingService
from services.user_service import UserService
# Removido: from services.report_service import ReportService
from services.segmento_atividade_service import SegmentoAtividadeService

# Tentar importar serviço completo, usar lite como fallback
try:
    from check_dependencies import get_available_import_service
    ImportService, IMPORT_SERVICE_TYPE = get_available_import_service()
    print(f"✅ Serviço de importação: {IMPORT_SERVICE_TYPE}")
except Exception as e:
    print(f"❌ Erro ao carregar serviço de importação: {e}")
    ImportService = None
    IMPORT_SERVICE_TYPE = "none"

from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production')

# Aplicar otimizações de memória RENDER-OTIMIZADAS se disponível
if MEMORY_OPTIMIZER_AVAILABLE:
    UltraMemoryOptimizer.setup_extreme_memory_optimization()
    UltraMemoryOptimizer.optimize_flask_config(app)
    
    # Configurações específicas do Render
    if os.environ.get('RENDER'):  # Detectar ambiente Render
        UltraMemoryOptimizer.setup_render_optimizations()
        print("🎯 Otimizações específicas Render aplicadas")
else:
    print("⚠️ Usando configurações básicas de memória")

# Configurações de encoding UTF-8
app.config['JSON_AS_ASCII'] = False
app.jinja_env.globals['ord'] = ord  # Função para debug de encoding
import sys
if hasattr(sys, 'setdefaultencoding'):
    sys.setdefaultencoding('utf-8')

# Forçar encoding UTF-8 para templates Jinja2
app.jinja_env.charset = 'utf-8'

# Filtro customizado para remover formatação de CPF/CNPJ
@app.template_filter('only_numbers')
def only_numbers_filter(value):
    """Remove toda formatação, deixando apenas números"""
    if not value:
        return ''
    import re
    return re.sub(r'\D', '', str(value))

# Filtro customizado para formatação de datas
@app.template_filter('format_date')
def format_date_filter(value, format='%d/%m/%Y'):
    """Formata data ISO para formato brasileiro"""
    if not value:
        return '-'
    try:
        # Se já for string no formato ISO
        if isinstance(value, str):
            # Remove microsegundos se existirem
            if '.' in value:
                value = value.split('.')[0]
            # Tenta parsing ISO
            if 'T' in value:
                dt = datetime.fromisoformat(value.replace('T', ' '))
            else:
                dt = datetime.strptime(value, '%Y-%m-%d')
        elif isinstance(value, datetime):
            dt = value
        else:
            return str(value)
        
        return dt.strftime(format)
    except (ValueError, AttributeError):
        return str(value)

# Filtro customizado para formatação de data e hora
@app.template_filter('format_datetime')
def format_datetime_filter(value):
    """Formata data e hora ISO para formato brasileiro"""
    return format_date_filter(value, '%d/%m/%Y %H:%M')

# Filtro específico para Data de Início dos Serviços - CORREÇÃO 05
@app.template_filter('format_mm_yyyy')
def format_mm_yyyy_filter(value):
    """Formata/valida data no formato MM/AAAA para Data de Início dos Serviços"""
    if not value:
        return '-'
    
    # Se já está no formato MM/AAAA, retorna como está
    if isinstance(value, str) and '/' in value:
        # Validar se está no formato correto MM/AAAA
        import re
        if re.match(r'^(0[1-9]|1[0-2])\/\d{4}$', value.strip()):
            return value.strip()
    
    # Se é uma string com apenas números, tenta converter para MM/AAAA
    if isinstance(value, str):
        digits_only = ''.join(filter(str.isdigit, value))
        if len(digits_only) == 6:  # MMAAAA
            month = digits_only[:2]
            year = digits_only[2:]
            if 1 <= int(month) <= 12:
                return f"{month}/{year}"
    
    # Retorna o valor original se não conseguir formatar
    return str(value) if value else '-'

# Configurações para upload de arquivos - OTIMIZADO PARA RENDER
# Reduzido DRASTICAMENTE para economizar RAM no Render (256KB)
MAX_UPLOAD_SIZE = 256 * 1024 if os.environ.get('RENDER') else 1 * 1024 * 1024  # 256KB Render, 1MB desenvolvimento
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE
app.config['UPLOAD_FOLDER'] = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

print(f"📁 Upload configurado: {MAX_UPLOAD_SIZE / 1024:.0f}KB máximo")

# Configurações RENDER-ESPECÍFICAS para produção - FOCO EM MEMÓRIA MÍNIMA
if os.environ.get('RENDER'):
    # Garbage collection EXTREMAMENTE agressivo para Render
    gc.set_threshold(10, 1, 1)  # Muito mais agressivo que antes
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Zero cache para economizar memória
    
    # Configurações JSON EXTREMAS
    app.config['JSON_SORT_KEYS'] = False
    app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
    app.config['JSONIFY_MIMETYPE'] = 'application/json'
    
    # Limitar DRASTICAMENTE workers e conexões para Render
    os.environ.setdefault('WEB_CONCURRENCY', '1')  # 1 worker APENAS
    os.environ.setdefault('WORKER_CONNECTIONS', '3')  # Extremamente reduzido
    os.environ.setdefault('WORKER_TIMEOUT', '15')  # Timeout muito baixo
    os.environ.setdefault('MAX_REQUESTS', '25')  # Restart worker mais frequentemente
    os.environ.setdefault('PRELOAD_APP', 'true')  # Preload para economizar memória
    
    # Configurações de sessão otimizadas para Render
    app.config['PERMANENT_SESSION_LIFETIME'] = 43200  # 12 horas (43200 segundos)
    app.config['SESSION_COOKIE_SECURE'] = True
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Melhor compatibilidade
    
    # Desabilitar TUDO que consome memória desnecessariamente
    app.config['PRESERVE_CONTEXT_ON_EXCEPTION'] = False
    app.config['EXPLAIN_TEMPLATE_LOADING'] = False
    app.config['PROPAGATE_EXCEPTIONS'] = None
    
    print("🎯 Configurações RENDER aplicadas - MEMÓRIA MÍNIMA")
elif os.environ.get('FLASK_ENV') == 'production':
    # Configurações genéricas de produção (não Render)
    gc.set_threshold(25, 1, 1)
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 30
    app.config['PERMANENT_SESSION_LIFETIME'] = 43200  # 12 horas também em produção
    print("🧠 Configurações genéricas de produção aplicadas")
else:
    # Configurações para desenvolvimento local
    app.config['PERMANENT_SESSION_LIFETIME'] = 43200  # 12 horas em desenvolvimento
    app.config['SESSION_COOKIE_SECURE'] = False  # HTTP permitido em desenvolvimento
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    print("🔧 Configurações de desenvolvimento aplicadas - Sessão 12h")

# Criar pasta de uploads se não existir
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Hook EXTREMO para limpeza de memória após cada requisição - RENDER OTIMIZADO
@app.after_request
def cleanup_memory_after_request(response):
    """Limpa memória EXTREMAMENTE após cada requisição - otimizado para Render"""
    if MEMORY_OPTIMIZER_AVAILABLE and os.environ.get('RENDER'):
        # Usar otimizador específico para Render
        UltraMemoryOptimizer.cleanup_after_request()
        
        # Limpeza adicional para ambientes críticos de memória (Render)
        try:
            # Forçar limpeza de cache Python interno
            if hasattr(sys, '_clear_type_cache'):
                sys._clear_type_cache()
            
            # Coleta de lixo múltipla para Render
            for _ in range(3):  # Mais agressivo no Render
                gc.collect()
                
            # Força limpeza de weakrefs
            import weakref
            weakref.getweakrefs(object())
        except:
            pass
    elif os.environ.get('FLASK_ENV') == 'production':
        # Limpeza básica para outros ambientes de produção
        for _ in range(2):
            gc.collect()
    
    return response

# Carregar variáveis de ambiente (.env local / Render)
from dotenv import load_dotenv
load_dotenv()  # Carrega .env apenas localmente (Render usa variáveis nativas)

# Carregar configurações (compatível com produção)
USE_GOOGLE_SHEETS = True
USE_OAUTH2 = False  # OAuth2 para autenticação manual  
USE_SERVICE_ACCOUNT = True  # Service Account para aplicações server-side (RECOMENDADO)
GOOGLE_SHEETS_API_KEY = os.environ.get('GOOGLE_SHEETS_API_KEY')
GOOGLE_SHEETS_ID = os.environ.get('GOOGLE_SHEETS_ID')
GOOGLE_SHEETS_RANGE = 'Clientes!A:DD'

print(f"🔧 Configurações:")
print(f"   USE_GOOGLE_SHEETS: {USE_GOOGLE_SHEETS}")
print(f"   USE_OAUTH2: {USE_OAUTH2}")
print(f"   USE_SERVICE_ACCOUNT: {USE_SERVICE_ACCOUNT}")
if GOOGLE_SHEETS_API_KEY:
    print(f"   API_KEY: {GOOGLE_SHEETS_API_KEY[:10]}...")
else:
    print("   API_KEY: None")
print(f"   SPREADSHEET_ID: {GOOGLE_SHEETS_ID}")

# Inicializar serviços COM LAZY LOADING - OTIMIZAÇÃO MEMÓRIA RENDER
storage_service = None
meeting_service = None
user_service = None
# Removido: report_service = None (funcionalidade de relatórios removida)
import_service = None

def get_storage_service():
    """Lazy loading do storage service com fallback robusto para Render"""
    global storage_service
    if storage_service is None:
        try:
            if USE_GOOGLE_SHEETS and GOOGLE_SHEETS_ID:
                if USE_SERVICE_ACCOUNT:
                    print("🔐 Inicializando Google Sheets Service Account...")
                    from services.google_sheets_service_account import GoogleSheetsServiceAccountService
                    storage_service = GoogleSheetsServiceAccountService(GOOGLE_SHEETS_ID, GOOGLE_SHEETS_RANGE)
                    print("✅ Storage service inicializado (Service Account)")
                elif USE_OAUTH2:
                    print("🔐 Inicializando Google Sheets OAuth2...")
                    from services.google_sheets_oauth_service import GoogleSheetsOAuthService
                    storage_service = GoogleSheetsOAuthService(GOOGLE_SHEETS_ID, GOOGLE_SHEETS_RANGE)
                    print("✅ Storage service inicializado (OAuth2)")
                else:
                    print("📊 Inicializando Google Sheets híbrido...")
                    storage_service = GoogleSheetsService(GOOGLE_SHEETS_API_KEY, GOOGLE_SHEETS_ID, GOOGLE_SHEETS_RANGE)
                    print("✅ Storage service inicializado (Híbrido)")
            else:
                print("⚠️ Usando armazenamento local")
                storage_service = LocalStorageService()
                
        except Exception as e:
            print(f"❌ Erro ao inicializar storage service: {e}")
            print(f"🔍 Tipo do erro: {type(e).__name__}")
            
            # Verificar se é erro de autenticação específico do Render
            error_str = str(e).lower()
            is_auth_error = any(keyword in error_str for keyword in [
                'authentication', 'credential', 'unauthorized', 'forbidden',
                'service account', 'google', 'api', 'quota', 'permission'
            ])
            
            if is_auth_error and os.environ.get('FLASK_ENV') == 'production':
                print("🚨 Erro de autenticação detectado no Render - ativando fallback")
                try:
                    from services.render_fallback_service import RenderFallbackService
                    storage_service = RenderFallbackService()
                    print("✅ Render Fallback Service ativado")
                    
                    # Adicionar mensagem global para mostrar na interface
                    if not hasattr(get_storage_service, '_fallback_message_shown'):
                        flash('⚠️ Sistema temporariamente usando dados locais. Verifique configurações do Google Sheets.', 'warning')
                        get_storage_service._fallback_message_shown = True
                        
                except Exception as fallback_error:
                    print(f"❌ Erro ao ativar fallback: {fallback_error}")
                    storage_service = LocalStorageService()
                    print("⚠️ Fallback final para armazenamento local")
            else:
                storage_service = LocalStorageService()
                print("⚠️ Fallback para armazenamento local")
        
        # Limpeza de memória após inicialização
        if MEMORY_OPTIMIZER_AVAILABLE:
            UltraMemoryOptimizer.force_cleanup()
            print(f"💾 Memória após limpeza completa: {UltraMemoryOptimizer.get_memory_usage()}")
    
    return storage_service

def get_meeting_service():
    """Lazy loading do meeting service"""
    global meeting_service
    if meeting_service is None and GOOGLE_SHEETS_ID:
        try:
            meeting_service = MeetingService(GOOGLE_SHEETS_ID)
            print("✅ Meeting service inicializado")
        except Exception as e:
            print(f"❌ Erro ao inicializar meeting service: {e}")
            meeting_service = None
    return meeting_service

def get_user_service():
    """Lazy loading do user service com fallback"""
    global user_service
    if user_service is None:
        if GOOGLE_SHEETS_ID:
            try:
                print(f"🔄 Tentando inicializar UserService com SHEETS_ID: {GOOGLE_SHEETS_ID}")
                user_service = UserService(GOOGLE_SHEETS_ID)
                print("✅ User service inicializado com sucesso")
            except Exception as e:
                print(f"❌ Erro ao inicializar user service: {e}")
                print(f"🔍 Tipo do erro: {type(e).__name__}")
                print("🔄 Tentando fallback de usuários...")
                
                # Usar serviço de fallback
                try:
                    from services.fallback_user_service import FallbackUserService
                    user_service = FallbackUserService()
                    print("✅ Fallback user service inicializado")
                except Exception as fallback_error:
                    print(f"❌ Erro no fallback user service: {fallback_error}")
                    user_service = None
        else:
            print("❌ GOOGLE_SHEETS_ID não disponível para UserService")
            print("🔄 Inicializando fallback user service...")
            
            # Usar serviço de fallback quando não há SHEETS_ID
            try:
                from services.fallback_user_service import FallbackUserService
                user_service = FallbackUserService()
                print("✅ Fallback user service inicializado (sem SHEETS_ID)")
            except Exception as fallback_error:
                print(f"❌ Erro no fallback user service: {fallback_error}")
                user_service = None
    else:
        print("♻️ User service já inicializado")
    return user_service

# Removido: Função get_report_service() (funcionalidade de relatórios removida)

def get_import_service():
    """Lazy loading do import service"""
    global import_service
    if import_service is None and ImportService:
        try:
            storage = get_storage_service()
            if storage:
                import_service = ImportService(storage)
                print(f"✅ Import service inicializado ({IMPORT_SERVICE_TYPE})")
        except Exception as e:
            print(f"❌ Erro ao inicializar import service: {e}")
            import_service = None
    return import_service

# Global para cache do serviço de segmentos/atividades
segmento_atividade_service = None

def get_segmento_atividade_service():
    """Lazy loading do segmento/atividade service"""
    global segmento_atividade_service
    if segmento_atividade_service is None:
        try:
            if GOOGLE_SHEETS_ID:
                print("🏢 Inicializando SegmentoAtividadeService...")
                segmento_atividade_service = SegmentoAtividadeService(GOOGLE_SHEETS_ID)
                print("✅ SegmentoAtividadeService inicializado")
            else:
                print("❌ GOOGLE_SHEETS_ID não disponível para SegmentoAtividadeService")
        except Exception as e:
            print(f"❌ Erro ao inicializar SegmentoAtividadeService: {e}")
            import traceback
            traceback.print_exc()
            segmento_atividade_service = None
    return segmento_atividade_service

# Inicialização básica - EXTREMAMENTE otimizada
print("🚀 Aplicação inicializada com lazy loading EXTREMO")
print(f"💾 Memória inicial: {UltraMemoryOptimizer.get_memory_usage() if MEMORY_OPTIMIZER_AVAILABLE else 'N/A'}")

# Iniciar monitor de memória ultra-otimizado
if os.environ.get('FLASK_ENV') == 'production' and MEMORY_OPTIMIZER_AVAILABLE:
    try:
        from ultra_memory_monitor import start_ultra_monitoring
        start_ultra_monitoring()
        print("🔍 Monitor de memória ULTRA iniciado")
    except ImportError:
        print("⚠️ Monitor ultra não disponível")

# Garbage collection EXTREMO inicial
if os.environ.get('FLASK_ENV') == 'production':
    if MEMORY_OPTIMIZER_AVAILABLE:
        UltraMemoryOptimizer.force_cleanup()
    else:
        gc.collect()
    print("🧠 Limpeza EXTREMA de memória inicial concluída")

# Decorator para verificar autenticação
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        print(f"🔐 LOGIN_REQUIRED: Verificando sessão para função {f.__name__}")
        print(f"🔐 Dados da sessão: {dict(session)}")
        if 'user_id' not in session:
            print("❌ user_id não encontrado na sessão - Redirecionando para login")
            flash('Você precisa fazer login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        print(f"✅ user_id encontrado: {session['user_id']} - Executando função {f.__name__}")
        return f(*args, **kwargs)
    return decorated_function

# Função para verificar se é administrador
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Você precisa fazer login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        
        # Usar get_user_service() em vez de user_service diretamente
        current_user_service = get_user_service()
        if current_user_service:
            user = current_user_service.get_user_by_id(session['user_id'])
            if not user or user.get('perfil', '').lower() != 'administrador':
                flash('Acesso negado. Apenas administradores podem acessar esta página.', 'danger')
                return redirect(url_for('index'))
        else:
            # Se não há user_service, verificar se é o admin de fallback
            if session.get('user_id') != 'admin-fallback':
                flash('Acesso negado. Sistema em modo de manutenção.', 'danger')
                return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

# Rota de boas-vindas (não protegida) - redireciona para login ou dashboard
@app.route('/welcome')
def welcome():
    """Página de boas-vindas - redireciona baseado no status de autenticação"""
    print("🏠 WELCOME: Verificando status de autenticação...")
    
    # Se já está logado, redirecionar para dashboard
    if 'user_id' in session and session.get('user_id'):
        print(f"🏠 WELCOME: Usuário já autenticado (ID: {session['user_id']}) - Redirecionando para dashboard")
        return redirect(url_for('index'))
    
    # Se não está logado, redirecionar para login
    print("🏠 WELCOME: Usuário não autenticado - Redirecionando para login")
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    print(f"🔐 LOGIN: Método da requisição: {request.method}")
    
    if request.method == 'POST':
        # Aceitar tanto username/password quanto usuario/senha
        username = request.form.get('username') or request.form.get('usuario')
        password = request.form.get('password') or request.form.get('senha')
        
        if not username or not password:
            flash('Por favor, preencha usuário e senha.', 'error')
            return render_template('login.html')
        
        print(f"🔐 LOGIN: Tentativa de login para usuário: {username}")
        
        # Debug das variáveis de ambiente críticas
        print(f"🔍 DEBUG: GOOGLE_SHEETS_ID = {GOOGLE_SHEETS_ID}")
        print(f"🔍 DEBUG: USE_GOOGLE_SHEETS = {USE_GOOGLE_SHEETS}")
        print(f"🔍 DEBUG: USE_SERVICE_ACCOUNT = {USE_SERVICE_ACCOUNT}")
        
        # Tentar inicializar o user_service
        current_user_service = get_user_service()
        print(f"🔐 LOGIN: user_service disponível: {current_user_service is not None}")
        
        if current_user_service:
            print("🔐 LOGIN: Chamando authenticate_user...")
            user = current_user_service.authenticate_user(username, password)
            print(f"🔐 LOGIN: Resultado da autenticação: {user}")
            
            if user:
                session['user_id'] = user['id']
                session['user_name'] = user['nome']
                session['user_perfil'] = user['perfil']
                session.permanent = True
                print(f"🔐 LOGIN: Sessão criada - user_id: {session['user_id']}")
                flash(f'Bem-vindo(a), {user["nome"]}!', 'success')
                print("🔐 LOGIN: Redirecionando para seleção de sistemas...")
                return redirect(url_for('system_selection'))
            else:
                print("❌ LOGIN: Falha na autenticação")
                flash('Usuário ou senha incorretos.', 'error')
        else:
            print("❌ LOGIN: user_service não disponível")
            flash('Serviço de autenticação indisponível. Tente novamente.', 'error')
    else:
        print("🔐 LOGIN: Exibindo formulário de login (GET)")
    
    return render_template('login.html')

@app.route('/system-selection')
@login_required
def system_selection():
    """Tela de seleção de sistemas após o login"""
    print(f"🎯 SYSTEM_SELECTION: Usuário {session.get('user_name')} acessando seleção de sistemas")
    return render_template('system_selection.html')

@app.route('/get-user-systems')
@login_required
def get_user_systems():
    """Retorna os sistemas disponíveis para o usuário baseado em suas permissões"""
    try:
        user_id = session.get('user_id')
        print(f"🎯 GET_USER_SYSTEMS: Buscando sistemas para usuário {user_id}")
        
        # Buscar dados do usuário para verificar permissões
        current_user_service = get_user_service()
        if current_user_service:
            user_data = current_user_service.get_user_by_id(user_id)
            if user_data:
                # Sistemas disponíveis baseado nas permissões do usuário
                available_systems = []
                
                # SIGEC sempre disponível (mínimo)
                available_systems.append('sigec')
                
                # Verificar se é administrador - tem acesso a tudo
                user_perfil = user_data.get('perfil', '').lower()
                if user_perfil == 'administrador':
                    # Administradores têm acesso a todos os sistemas
                    available_systems = ['sigec', 'operacao-fiscal', 'gestao-operacional', 'gestao-financeira']
                    print(f"🔑 GET_USER_SYSTEMS: Usuário administrador - todos os sistemas disponíveis")
                    
                    return {
                        'success': True,
                        'systems': available_systems,
                        'user_permissions': 'TOTAL_CADASTROS'  # Administradores têm permissão total
                    }
                
                # Para usuários não-administradores, verificar permissões específicas
                user_systems = user_data.get('sistemas_acesso', 'sigec')
                if isinstance(user_systems, str):
                    user_systems = [s.strip() for s in user_systems.split(',') if s.strip()]
                
                for system in user_systems:
                    system_lower = system.lower().strip()
                    if system_lower == 'operacao-fiscal':
                        available_systems.append('operacao-fiscal')
                    elif system_lower == 'gestao-operacional':
                        available_systems.append('gestao-operacional')
                    elif system_lower == 'gestao-financeira':
                        available_systems.append('gestao-financeira')
                
                # Remover duplicatas mantendo ordem
                available_systems = list(dict.fromkeys(available_systems))
                
                print(f"🎯 GET_USER_SYSTEMS: Sistemas disponíveis: {available_systems}")
                
                return {
                    'success': True,
                    'systems': available_systems,
                    'user_permissions': user_data.get('permissoes_sigec', 'VISUALIZADOR')
                }
            else:
                print("❌ GET_USER_SYSTEMS: Usuário não encontrado")
                return {'success': False, 'message': 'Usuário não encontrado'}, 404
        else:
            print("❌ GET_USER_SYSTEMS: Serviço de usuário indisponível")
            # Fallback: retornar apenas SIGEC
            return {
                'success': True,
                'systems': ['sigec'],
                'user_permissions': 'VISUALIZADOR'
            }
            
    except Exception as e:
        print(f"❌ GET_USER_SYSTEMS: Erro: {str(e)}")
        return {'success': False, 'message': 'Erro interno do servidor'}, 500

@app.route('/select-system', methods=['POST'])
@login_required
def select_system():
    """Processa a seleção do sistema e redireciona para o sistema escolhido"""
    try:
        data = request.get_json()
        system_type = data.get('system')
        
        print(f"🎯 SELECT_SYSTEM: Usuário {session.get('user_name')} selecionou sistema: {system_type}")
        
        # Armazenar o sistema selecionado na sessão
        session['selected_system'] = system_type
        
        # Definir URLs de redirecionamento baseado no sistema
        redirect_urls = {
            'sigec': url_for('index'),  # Dashboard principal atual
            'operacao-fiscal': '/operacao-fiscal',  # Sistema fiscal (placeholder)
            'gestao-operacional': 'https://app.powerbi.com/reportEmbed?reportId=8165cd63-42f4-44c1-8e4a-cad1a32d0e5b&autoAuth=true&ctid=0b754a09-0568-48fd-a100-8621a0bbd7ab',  # Power BI Gestão Operacional
            'gestao-financeira': 'https://app.powerbi.com/reportEmbed?reportId=ef9c9663-7cec-4c9a-8b57-c1a6c895057a&autoAuth=true&ctid=0b754a09-0568-48fd-a100-8621a0bbd7ab'  # Power BI Gestão Financeira
        }
        
        redirect_url = redirect_urls.get(system_type, url_for('index'))
        
        return {
            'success': True,
            'redirect_url': redirect_url,
            'message': f'Sistema {system_type} selecionado com sucesso!'
        }
        
    except Exception as e:
        print(f"❌ SELECT_SYSTEM: Erro ao processar seleção: {str(e)}")
        return {
            'success': False,
            'message': 'Erro interno do servidor'
        }, 500

@app.route('/logout')
def logout():
    session.clear()
    flash('Logout realizado com sucesso.', 'info')
    return redirect(url_for('login'))

# === ROTAS PARA SISTEMAS ESPECÍFICOS ===
@app.route('/operacao-fiscal')
@login_required
def operacao_fiscal():
    """Sistema de Operação Fiscal - Placeholder"""
    flash('Sistema de Operação Fiscal em desenvolvimento.', 'info')
    return render_template('under_construction.html', 
                         system_name='Operação Fiscal',
                         description='Sistema de controle fiscal e tributário')

# Removido: Rotas /gestao-operacional e /gestao-financeira (funcionalidade Power BI removida)

# === FUNÇÕES AUXILIARES PARA SEGMENTOS E ATIVIDADES ===
def get_segmentos_list():
    """Retorna lista de segmentos cadastrados"""
    try:
        service = get_segmento_atividade_service()
        if service:
            segmentos = service.get_segmentos_ativos()
            return segmentos  # Retorna objetos completos com id e nome
        else:
            print("⚠️ Serviço não disponível, usando lista padrão")
    except Exception as e:
        print(f"❌ Erro ao buscar segmentos: {e}")
    
    # Fallback para lista estática
    return [
        {"id": 1, "nome": "COMÉRCIO VAREJISTA"},
        {"id": 2, "nome": "COMÉRCIO ATACADISTA"},
        {"id": 3, "nome": "PRESTAÇÃO DE SERVIÇOS"},
        {"id": 4, "nome": "INDÚSTRIA"},
        {"id": 5, "nome": "AGRONEGÓCIO"},
        {"id": 6, "nome": "TECNOLOGIA"},
        {"id": 7, "nome": "EDUCAÇÃO"},
        {"id": 8, "nome": "SAÚDE"},
        {"id": 9, "nome": "ALIMENTAÇÃO"},
        {"id": 10, "nome": "OUTROS"}
    ]

def get_atividades_list():
    """Retorna lista de atividades principais cadastradas"""
    try:
        service = get_segmento_atividade_service()
        if service:
            atividades = service.get_atividades_ativas()
            return atividades  # Retorna objetos completos com id e nome
        else:
            print("⚠️ Serviço não disponível, usando lista padrão")
    except Exception as e:
        print(f"❌ Erro ao buscar atividades: {e}")
    
    # Fallback para lista estática
    return [
        {"id": 1, "nome": "VENDA DE ROUPAS E ACESSÓRIOS"},
        {"id": 2, "nome": "CONSULTORIA EMPRESARIAL"},
        {"id": 3, "nome": "DESENVOLVIMENTO DE SOFTWARE"},
        {"id": 4, "nome": "SERVIÇOS CONTÁBEIS"},
        {"id": 5, "nome": "RESTAURANTE E LANCHONETE"},
        {"id": 6, "nome": "CLÍNICA MÉDICA"},
        {"id": 7, "nome": "ESCOLA DE ENSINO FUNDAMENTAL"},
        {"id": 8, "nome": "TRANSPORTE RODOVIÁRIO DE CARGAS"},
        {"id": 9, "nome": "CONSTRUÇÃO DE EDIFÍCIOS"},
        {"id": 10, "nome": "PRODUÇÃO AGRÍCOLA"}
    ]

def get_sistemas_list():
    """Retorna lista de sistemas utilizados disponíveis"""
    # Lista estática de sistemas mais utilizados
    return [
        {"id": 1, "nome": "FORTES"},
        {"id": 2, "nome": "DOMÍNIO"},
        {"id": 3, "nome": "SAGE"},
        {"id": 4, "nome": "ALTERDATA"},
        {"id": 5, "nome": "ECONET"},
        {"id": 6, "nome": "SISPAG"},
        {"id": 7, "nome": "FOLHAMATIC"},
        {"id": 8, "nome": "PROSOFT"},
        {"id": 9, "nome": "SISTEMA PRÓPRIO"},
        {"id": 10, "nome": "PLANILHAS EXCEL"},
        {"id": 11, "nome": "SISTEMA ONLINE"},
        {"id": 12, "nome": "OUTROS"}
    ]

# === ROTAS PARA CADASTRO DE SEGMENTOS E ATIVIDADES ===
@app.route('/segmentos')
@login_required
def manage_segmentos():
    """Página para gerenciar segmentos"""
    try:
        service = get_segmento_atividade_service()
        if service:
            segmentos = service.get_segmentos()
            return render_template('manage_segmentos.html', segmentos=segmentos)
        else:
            flash('Serviço de segmentos indisponível', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        print(f"❌ Erro ao carregar segmentos: {e}")
        flash('Erro ao carregar segmentos', 'error')
        return redirect(url_for('index'))

@app.route('/segmentos/create', methods=['POST'])
@login_required  
def create_segmento():
    """Criar novo segmento"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            flash('Serviço de segmentos indisponível', 'error')
            return redirect(url_for('manage_segmentos'))
        
        nome_segmento = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        codigo = request.form.get('codigo', '').strip()
        
        if not nome_segmento:
            flash('Nome do segmento é obrigatório', 'error')
            return redirect(url_for('manage_segmentos'))
        
        segmento_data = {
            'nome': nome_segmento.upper(),
            'descricao': descricao,
            'codigo': codigo.upper() if codigo else '',
            'ativo': True,
            'criadoPor': session.get('user_name', 'Sistema')
        }
        
        if service.save_segmento(segmento_data):
            flash(f'Segmento "{nome_segmento}" criado com sucesso!', 'success')
        else:
            flash('Erro ao criar segmento', 'error')
            
        return redirect(url_for('manage_segmentos'))
    except Exception as e:
        print(f"❌ Erro ao criar segmento: {e}")
        flash('Erro ao criar segmento', 'error')
        return redirect(url_for('manage_segmentos'))

@app.route('/segmentos/edit/<segmento_id>', methods=['POST'])
@login_required
def edit_segmento(segmento_id):
    """Editar segmento existente"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            flash('Serviço de segmentos indisponível', 'error')
            return redirect(url_for('manage_segmentos'))
        
        segmento = service.get_segmento(segmento_id)
        if not segmento:
            flash('Segmento não encontrado', 'error')
            return redirect(url_for('manage_segmentos'))
        
        # Atualizar dados
        segmento['nome'] = request.form.get('nome', '').strip().upper()
        segmento['descricao'] = request.form.get('descricao', '').strip()
        segmento['codigo'] = request.form.get('codigo', '').strip().upper()
        segmento['ativo'] = request.form.get('ativo') == 'on'
        
        if not segmento['nome']:
            flash('Nome do segmento é obrigatório', 'error')
            return redirect(url_for('manage_segmentos'))
        
        if service.save_segmento(segmento):
            flash(f'Segmento "{segmento["nome"]}" atualizado com sucesso!', 'success')
        else:
            flash('Erro ao atualizar segmento', 'error')
            
        return redirect(url_for('manage_segmentos'))
    except Exception as e:
        print(f"❌ Erro ao editar segmento: {e}")
        flash('Erro ao editar segmento', 'error')
        return redirect(url_for('manage_segmentos'))

@app.route('/segmentos/delete/<segmento_id>', methods=['POST'])
@login_required
def delete_segmento(segmento_id):
    """Excluir segmento (soft delete)"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            flash('Serviço de segmentos indisponível', 'error')
            return redirect(url_for('manage_segmentos'))
        
        if service.delete_segmento(segmento_id):
            flash('Segmento excluído com sucesso!', 'success')
        else:
            flash('Erro ao excluir segmento', 'error')
            
        return redirect(url_for('manage_segmentos'))
    except Exception as e:
        print(f"❌ Erro ao excluir segmento: {e}")
        flash('Erro ao excluir segmento', 'error')
        return redirect(url_for('manage_segmentos'))

@app.route('/atividades')
@login_required
def manage_atividades():
    """Página para gerenciar atividades principais"""
    try:
        service = get_segmento_atividade_service()
        if service:
            atividades = service.get_atividades()
            segmentos = service.get_segmentos_ativos()
            return render_template('manage_atividades.html', atividades=atividades, segmentos=segmentos)
        else:
            flash('Serviço de atividades indisponível', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        print(f"❌ Erro ao carregar atividades: {e}")
        flash('Erro ao carregar atividades', 'error')
        return redirect(url_for('index'))

@app.route('/atividades/create', methods=['POST'])
@login_required
def create_atividade():
    """Criar nova atividade principal"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            flash('Serviço de atividades indisponível', 'error')
            return redirect(url_for('manage_atividades'))
        
        nome_atividade = request.form.get('nome', '').strip()
        codigo_cnae = request.form.get('codigoCnae', '').strip()
        descricao = request.form.get('descricao', '').strip()
        segmento_id = request.form.get('segmentoId', '').strip()
        
        if not nome_atividade:
            flash('Nome da atividade é obrigatório', 'error')
            return redirect(url_for('manage_atividades'))
        
        if not segmento_id:
            flash('Segmento é obrigatório', 'error')
            return redirect(url_for('manage_atividades'))
        
        atividade_data = {
            'nome': nome_atividade.upper(),
            'codigoCnae': codigo_cnae,
            'descricao': descricao,
            'segmentoId': segmento_id,
            'ativo': True,
            'criadoPor': session.get('user_name', 'Sistema')
        }
        
        if service.save_atividade(atividade_data):
            flash(f'Atividade "{nome_atividade}" criada com sucesso!', 'success')
        else:
            flash('Erro ao criar atividade', 'error')
            
        return redirect(url_for('manage_atividades'))
    except Exception as e:
        print(f"❌ Erro ao criar atividade: {e}")
        flash('Erro ao criar atividade', 'error')
        return redirect(url_for('manage_atividades'))

@app.route('/atividades/edit/<atividade_id>', methods=['POST'])
@login_required
def edit_atividade(atividade_id):
    """Editar atividade existente"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            flash('Serviço de atividades indisponível', 'error')
            return redirect(url_for('manage_atividades'))
        
        atividade = service.get_atividade(atividade_id)
        if not atividade:
            flash('Atividade não encontrada', 'error')
            return redirect(url_for('manage_atividades'))
        
        # Atualizar dados
        atividade['nome'] = request.form.get('nome', '').strip().upper()
        atividade['codigoCnae'] = request.form.get('codigoCnae', '').strip()
        atividade['descricao'] = request.form.get('descricao', '').strip()
        atividade['segmentoId'] = request.form.get('segmentoId', '').strip()
        atividade['ativo'] = request.form.get('ativo') == 'on'
        
        if not atividade['nome']:
            flash('Nome da atividade é obrigatório', 'error')
            return redirect(url_for('manage_atividades'))
        
        if not atividade['segmentoId']:
            flash('Segmento é obrigatório', 'error')
            return redirect(url_for('manage_atividades'))
        
        if service.save_atividade(atividade):
            flash(f'Atividade "{atividade["nome"]}" atualizada com sucesso!', 'success')
        else:
            flash('Erro ao atualizar atividade', 'error')
            
        return redirect(url_for('manage_atividades'))
    except Exception as e:
        print(f"❌ Erro ao editar atividade: {e}")
        flash('Erro ao editar atividade', 'error')
        return redirect(url_for('manage_atividades'))

@app.route('/atividades/delete/<atividade_id>', methods=['POST'])
@login_required
def delete_atividade(atividade_id):
    """Excluir atividade (soft delete)"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            flash('Serviço de atividades indisponível', 'error')
            return redirect(url_for('manage_atividades'))
        
        if service.delete_atividade(atividade_id):
            flash('Atividade excluída com sucesso!', 'success')
        else:
            flash('Erro ao excluir atividade', 'error')
            
        return redirect(url_for('manage_atividades'))
    except Exception as e:
        print(f"❌ Erro ao excluir atividade: {e}")
        flash('Erro ao excluir atividade', 'error')
        return redirect(url_for('manage_atividades'))

# === ROTAS API PARA INTEGRAÇÃO ===

@app.route('/api/segmentos', methods=['GET'])
@login_required
def api_get_segmentos():
    """API para buscar segmentos"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            return jsonify({'error': 'Serviço indisponível'}), 500
        
        apenas_ativos = request.args.get('ativos', 'true').lower() == 'true'
        termo_busca = request.args.get('busca', '').strip()
        
        if termo_busca:
            segmentos = service.search_segmentos(termo_busca)
        elif apenas_ativos:
            segmentos = service.get_segmentos_ativos()
        else:
            segmentos = service.get_segmentos()
        
        return jsonify({
            'success': True,
            'data': segmentos,
            'total': len(segmentos)
        })
    except Exception as e:
        print(f"❌ Erro na API de segmentos: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/atividades', methods=['GET'])
@login_required
def api_get_atividades():
    """API para buscar atividades"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            return jsonify({'error': 'Serviço indisponível'}), 500
        
        segmento_id = request.args.get('segmento_id')
        apenas_ativos = request.args.get('ativos', 'true').lower() == 'true'
        termo_busca = request.args.get('busca', '').strip()
        
        if termo_busca:
            atividades = service.search_atividades(termo_busca, segmento_id)
        elif apenas_ativos:
            atividades = service.get_atividades_ativas(segmento_id)
        else:
            atividades = service.get_atividades(segmento_id)
        
        return jsonify({
            'success': True,
            'data': atividades,
            'total': len(atividades)
        })
    except Exception as e:
        print(f"❌ Erro na API de atividades: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/segmentos/<segmento_id>', methods=['GET'])
@login_required
def api_get_segmento(segmento_id):
    """API para buscar um segmento específico"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            return jsonify({'error': 'Serviço indisponível'}), 500
        
        segmento = service.get_segmento(segmento_id)
        if not segmento:
            return jsonify({'error': 'Segmento não encontrado'}), 404
        
        return jsonify({
            'success': True,
            'data': segmento
        })
    except Exception as e:
        print(f"❌ Erro na API de segmento: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/atividades/<atividade_id>', methods=['GET'])
@login_required
def api_get_atividade(atividade_id):
    """API para buscar uma atividade específica"""
    try:
        service = get_segmento_atividade_service()
        if not service:
            return jsonify({'error': 'Serviço indisponível'}), 500
        
        atividade = service.get_atividade(atividade_id)
        if not atividade:
            return jsonify({'error': 'Atividade não encontrada'}), 404
        
        return jsonify({
            'success': True,
            'data': atividade
        })
    except Exception as e:
        print(f"❌ Erro na API de atividade: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/users')
@admin_required
def users():
    current_user_service = get_user_service()
    if current_user_service:
        users_list = current_user_service.list_users()
        return render_template('users.html', users=users_list)
    else:
        flash('Serviço de usuários indisponível.', 'error')
        return redirect(url_for('index'))

@app.route('/create_user', methods=['POST'])
@admin_required
def create_user():
    current_user_service = get_user_service()
    if current_user_service:
        nome = request.form['nome']
        email = request.form['email']
        usuario = request.form['usuario']
        senha = request.form['senha']
        perfil = request.form['perfil']
        
        # Novos campos para sistemas e permissões
        sistemas_acesso = request.form.getlist('sistemas_acesso')  # Lista de sistemas selecionados
        if not sistemas_acesso:
            sistemas_acesso = ['sigec']  # SIGEC sempre disponível
        sistemas_str = ','.join(sistemas_acesso)
        
        permissoes_sigec = request.form.get('permissoes_sigec', 'VISUALIZADOR')
        
        result = current_user_service.create_user(nome, email, usuario, senha, perfil, sistemas_str, permissoes_sigec)
        
        if result['success']:
            flash(result['message'], 'success')
        else:
            flash(result['message'], 'error')
    else:
        flash('Serviço de usuários indisponível.', 'error')
    
    return redirect(url_for('users'))

@app.route('/edit_user', methods=['POST'])
@admin_required
def edit_user():
    current_user_service = get_user_service()
    if current_user_service:
        user_id = request.form['user_id']
        nome = request.form['nome']
        email = request.form['email']
        usuario = request.form['usuario']
        perfil = request.form['perfil']
        ativo = request.form['ativo']
        nova_senha = request.form.get('nova_senha', '').strip()
        
        # Novos campos para sistemas e permissões
        sistemas_acesso = request.form.getlist('sistemas_acesso')
        if not sistemas_acesso:
            sistemas_acesso = ['sigec']  # SIGEC sempre disponível
        sistemas_str = ','.join(sistemas_acesso)
        
        permissoes_sigec = request.form.get('permissoes_sigec', 'VISUALIZADOR')
        
        # Se nova senha foi fornecida, usa ela, senão None
        senha_param = nova_senha if nova_senha else None
        
        result = current_user_service.update_user(user_id, nome, email, usuario, perfil, ativo, senha_param, sistemas_str, permissoes_sigec)
        
        if result['success']:
            flash(result['message'], 'success')
        else:
            flash(result['message'], 'error')
    else:
        flash('Serviço de usuários indisponível.', 'error')
    
    return redirect(url_for('users'))

@app.route('/delete_user', methods=['POST'])
@admin_required
def delete_user():
    current_user_service = get_user_service()
    if current_user_service:
        user_id = request.form['user_id']
        
        result = current_user_service.delete_user(user_id)
        
        if result['success']:
            flash(result['message'], 'success')
        else:
            flash(result['message'], 'error')
    else:
        flash('Serviço de usuários indisponível.', 'error')
    
    return redirect(url_for('users'))

@app.route('/import')
@admin_required
def import_page():
    """Página para importar clientes de planilha Excel"""
    return render_template('import.html')

@app.route('/import/upload', methods=['POST'])
@admin_required
def upload_and_import():
    """Processa upload e importa arquivo Excel com logs detalhados"""
    print("📤 === INICIANDO UPLOAD E IMPORTAÇÃO APRIMORADA ===")
    
    import_logs = []
    import_stats = {'success': 0, 'errors': 0, 'total': 0, 'details': []}
    
    # Obter serviço de importação
    import_service = get_import_service()
    print(f"🔧 Import service obtido: {import_service is not None}")
    if import_service:
        print(f"✅ Serviço disponível: {import_service.is_available()}")
    
    def add_log(message, level='info'):
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_entry = {'timestamp': timestamp, 'message': message, 'level': level}
        import_logs.append(log_entry)
        print(f"[{timestamp}] {level.upper()}: {message}")
    
    try:
        add_log("Iniciando processo de importação", "info")
        
        # Verificar se arquivo foi enviado
        if 'file' not in request.files:
            add_log("Erro: Nenhum arquivo foi enviado na requisição", "error")
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(url_for('import_page'))
        
        file = request.files['file']
        add_log(f"Arquivo recebido: {file.filename}", "info")
        
        # Verificar se arquivo tem nome
        if file.filename == '':
            add_log("Erro: Nome do arquivo está vazio", "error")
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(url_for('import_page'))
        
        # Verificar extensão do arquivo
        if not allowed_file(file.filename):
            add_log(f"Erro: Extensão inválida para arquivo {file.filename}", "error")
            flash('Apenas arquivos .xlsx e .xls são permitidos', 'error')
            return redirect(url_for('import_page'))
        
        add_log("Validação inicial do arquivo: APROVADA", "success")
        
        # Salvar arquivo temporariamente
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        add_log(f"Salvando arquivo temporário: {filename}", "info")
        file.save(file_path)
        
        file_size = os.path.getsize(file_path)
        add_log(f"Arquivo salvo com sucesso ({file_size} bytes)", "success")
        
        # Validar estrutura do arquivo
        if import_service and import_service.is_available():
            add_log("Serviço de importação disponível", "success")
            add_log("Iniciando validação da estrutura do arquivo", "info")
            
            try:
                is_valid, validation_message = import_service.validate_excel_structure(file_path)
                
                if not is_valid:
                    add_log(f"Validação falhou: {validation_message}", "error")
                    # Remover arquivo
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        add_log("Arquivo temporário removido", "info")
                    
                    flash(f'Estrutura do arquivo inválida: {validation_message}', 'error')
                    return redirect(url_for('import_page'))
                
                add_log("Estrutura do arquivo: VÁLIDA", "success")
                add_log("Iniciando processo de importação dos dados", "info")
                
                # Executar importação
                sucessos, erros, lista_erros = import_service.import_from_excel(file_path)
                
                import_stats['success'] = sucessos
                import_stats['errors'] = erros
                import_stats['total'] = sucessos + erros
                import_stats['details'] = lista_erros
                
                add_log(f"Importação concluída: {sucessos} sucessos, {erros} erros", "info")
                
                # Remover arquivo temporário
                if os.path.exists(file_path):
                    os.remove(file_path)
                    add_log("Arquivo temporário removido", "info")
                
                # Mostrar resultados detalhados
                if sucessos > 0:
                    add_log(f"Clientes importados com sucesso: {sucessos}", "success")
                    flash(f'✅ Importação concluída: {sucessos} clientes importados com sucesso!', 'success')
                
                if erros > 0:
                    add_log(f"Erros encontrados durante importação: {erros}", "warning")
                    flash(f'⚠️ {erros} erro(s) encontrado(s) - veja detalhes abaixo', 'warning')
                    
                    # Mostrar detalhes dos erros
                    for i, erro in enumerate(lista_erros[:10]):  # Limitar a 10 erros
                        add_log(f"Erro {i+1}: {erro}", "error")
                        flash(f'❌ {erro}', 'error')
                    
                    if len(lista_erros) > 10:
                        flash(f'... e mais {len(lista_erros) - 10} erro(s)', 'error')
                        add_log(f"Total de {len(lista_erros)} erros (mostrando apenas os primeiros 10)", "warning")
                
                if sucessos == 0 and erros == 0:
                    add_log("Nenhum cliente foi processado - arquivo pode estar vazio", "warning")
                    flash('Nenhum cliente foi processado', 'warning')
                
                # Log de resumo final
                add_log("=== RESUMO DA IMPORTAÇÃO ===", "info")
                add_log(f"Total de registros processados: {import_stats['total']}", "info")
                add_log(f"Sucessos: {import_stats['success']}", "success")
                add_log(f"Erros: {import_stats['errors']}", "error" if import_stats['errors'] > 0 else "info")
                
            except Exception as import_error:
                add_log(f"Erro durante validação/importação: {str(import_error)}", "error")
                # Remover arquivo
                if os.path.exists(file_path):
                    os.remove(file_path)
                    add_log("Arquivo temporário removido após erro", "info")
                
                flash(f'Erro durante importação: {str(import_error)}', 'error')
                return redirect(url_for('import_page'))
        
        else:
            add_log("Serviço de importação não disponível - dependências faltando", "error")
            # Remover arquivo
            if os.path.exists(file_path):
                os.remove(file_path)
                add_log("Arquivo temporário removido", "info")
            
            flash('Serviço de importação não disponível. Pandas/OpenPyXL não estão instalados.', 'error')
            return redirect(url_for('import_page'))
    
    except Exception as e:
        add_log(f"Erro crítico durante importação: {str(e)}", "error")
        print(f"❌ Erro crítico: {e}")
        flash(f'Erro crítico durante importação: {str(e)}', 'error')
        
        # Limpar arquivo se houver erro
        try:
            if 'file_path' in locals() and os.path.exists(file_path):
                os.remove(file_path)
                add_log("Arquivo temporário removido após erro crítico", "info")
        except Exception as cleanup_error:
            add_log(f"Erro ao limpar arquivo temporário: {str(cleanup_error)}", "warning")
    
    finally:
        # Log final
        add_log("Processo de importação finalizado", "info")
        print("📤 === IMPORTAÇÃO FINALIZADA ===")
        
        # Em um ambiente real, você poderia salvar os logs em sessão ou banco de dados
        # para mostrar na interface posteriormente
    
    return redirect(url_for('index'))
    
    return redirect(url_for('index'))

@app.route('/import/template')
@admin_required  
def download_template():
    """Baixa template Excel para importação"""
    try:
        # Verificar se openpyxl está disponível
        try:
            from openpyxl import Workbook
            import io
        except ImportError as e:
            print(f"❌ OpenPyXL não encontrado: {e}")
            # Fallback: criar template CSV
            return create_simple_template()
        
        print("✅ OpenPyXL importado com sucesso")
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Cabeçalhos SIGEC - 8 Blocos Organizados
        headers = [
            # Bloco 1: Informações da Pessoa Jurídica (13 campos obrigatórios)
            'NOME DA EMPRESA', 'RAZÃO SOCIAL NA RECEITA', 'NOME FANTASIA NA RECEITA', 
            'CNPJ', 'PERFIL', 'INSCRIÇÃO ESTADUAL', 'INSCRIÇÃO MUNICIPAL', 
            'ESTADO', 'CIDADE', 'REGIME FEDERAL', 'REGIME ESTADUAL', 'SEGMENTO', 'ATIVIDADE',
            
            # Bloco 2: Serviços Prestados pela Control (12 campos)
            'SERVIÇO CT', 'SERVIÇO FS', 'SERVIÇO DP', 'SERVIÇO BPO FINANCEIRO', 
            'RESPONSÁVEL PELOS SERVIÇOS', 'DATA INÍCIO DOS SERVIÇOS',
            'CÓDIGO FORTES CT', 'CÓDIGO FORTES FS', 'CÓDIGO FORTES PS', 
            'CÓDIGO DOMÍNIO', 'SISTEMA UTILIZADO', 'MÓDULO SPED TRIER',
            
            # Bloco 3: Quadro Societário (6 campos)
            'SÓCIO 1 NOME', 'SÓCIO 1 CPF', 'SÓCIO 1 DATA NASCIMENTO', 
            'SÓCIO 1 ADMINISTRADOR', 'SÓCIO 1 COTAS', 'SÓCIO 1 RESPONSÁVEL LEGAL',
            
            # Bloco 4: Contatos (10 campos básicos + contatos detalhados)
            'TELEFONE FIXO', 'TELEFONE CELULAR', 'WHATSAPP', 
            'EMAIL PRINCIPAL', 'EMAIL SECUNDÁRIO', 'RESPONSÁVEL IMEDIATO',
            'EMAILS DOS SÓCIOS', 'CONTATO CONTADOR', 'TELEFONE CONTADOR', 'EMAIL CONTADOR',
            
            # Contatos Detalhados (até 10 contatos com 4 campos cada = 40 campos)
            'CONTATO_1_NOME', 'CONTATO_1_CARGO', 'CONTATO_1_TELEFONE', 'CONTATO_1_EMAIL',
            'CONTATO_2_NOME', 'CONTATO_2_CARGO', 'CONTATO_2_TELEFONE', 'CONTATO_2_EMAIL',
            'CONTATO_3_NOME', 'CONTATO_3_CARGO', 'CONTATO_3_TELEFONE', 'CONTATO_3_EMAIL',
            'CONTATO_4_NOME', 'CONTATO_4_CARGO', 'CONTATO_4_TELEFONE', 'CONTATO_4_EMAIL',
            'CONTATO_5_NOME', 'CONTATO_5_CARGO', 'CONTATO_5_TELEFONE', 'CONTATO_5_EMAIL',
            
            # Bloco 5: Sistemas e Acessos (7 campos)
            'SISTEMA PRINCIPAL', 'VERSÃO DO SISTEMA', 'CÓDIGO ACESSO SIMPLES NACIONAL',
            'CPF/CNPJ PARA ACESSO', 'PORTAL CLIENTE ATIVO', 'INTEGRAÇÃO DOMÍNIO', 'SISTEMA ONVIO',
            
            # Bloco 6: Senhas e Credenciais (20 campos)
            'ACESSO ISS', 'SENHA ISS', 'ACESSO SEFIN', 'SENHA SEFIN', 
            'ACESSO SEUMA', 'SENHA SEUMA', 'ACESSO EMPWEB', 'SENHA EMPWEB',
            'ACESSO FAP/INSS', 'SENHA FAP/INSS', 'ACESSO CRF', 'SENHA CRF',
            'EMAIL GESTOR', 'SENHA EMAIL GESTOR', 'ANVISA GESTOR', 'ANVISA EMPRESA',
            'ACESSO IBAMA', 'SENHA IBAMA', 'ACESSO SEMACE', 'SENHA SEMACE',
            
            # Bloco 7: Procurações (12 campos)
            'PROCURAÇÃO RFB', 'DATA PROCURAÇÃO RFB', 'PROCURAÇÃO RECEITA ESTADUAL', 'DATA PROCURAÇÃO RC',
            'PROCURAÇÃO CAIXA ECONÔMICA', 'DATA PROCURAÇÃO CX', 'PROCURAÇÃO PREVIDÊNCIA SOCIAL', 'DATA PROCURAÇÃO SW',
            'PROCURAÇÃO MUNICIPAL', 'DATA PROCURAÇÃO MUNICIPAL', 'OUTRAS PROCURAÇÕES', 'OBSERVAÇÕES PROCURAÇÕES',
            
            # Bloco 8: Observações e Dados Adicionais (12 campos)
            'OBSERVAÇÕES GERAIS', 'TAREFAS VINCULADAS', 'STATUS DO CLIENTE',
            'ÚLTIMA ATUALIZAÇÃO', 'RESPONSÁVEL ATUALIZAÇÃO', 'PRIORIDADE',
            'TAGS/CATEGORIAS', 'HISTÓRICO DE ALTERAÇÕES'
        ]
        
        # Dados de exemplo SIGEC
        example_data = [
            # Bloco 1: Informações da Pessoa Jurídica
            'EMPRESA EXEMPLO LTDA', 'EMPRESA EXEMPLO LTDA', 'Exemplo Empresa', 
            '12.345.678/0001-99', 'LUCRO PRESUMIDO', '123456789', '987654321', 
            'CE', 'FORTALEZA', 'LUCRO PRESUMIDO', 'NORMAL', 'COMÉRCIO', 'VENDA DE PRODUTOS',
            
            # Bloco 2: Serviços Prestados pela Control
            'SIM', 'SIM', 'NÃO', 'SIM', 
            'João da Silva', '2024-01-01',
            '12345', '67890', '', 
            'DOM123', 'FORTES', 'COMPLETO',
            
            # Bloco 3: Quadro Societário
            'JOÃO DA SILVA', '123.456.789-00', '1980-01-01', 
            'SIM', '100%', 'SIM',
            
            # Bloco 4: Contatos
            '(85) 3333-4444', '(85) 99999-8888', '(85) 99999-8888', 
            'teste@empresa.com', 'contato@empresa.com', 'Maria Santos',
            'joao@empresa.com', 'Contador ABC', '(85) 3333-5555', 'contador@abc.com',
            
            # Bloco 5: Sistemas e Acessos
            'FORTES', '2024.1', 'SN123456',
            '12345678000190', 'SIM', 'SIM', 'NÃO',
            
            # Bloco 6: Senhas e Credenciais
            'usuario123', 'senha123', 'sefin123', 'senha456', 
            'seuma123', 'senha789', 'empweb123', 'senha999',
            'fap123', 'senha000', 'crf123', 'senha111',
            'gestor@empresa.com', 'senha222', 'anvisa123', 'empresa123',
            'ibama123', 'senha333', 'semace123', 'senha444',
            
            # Bloco 7: Procurações
            'SIM', '2024-01-15', 'SIM', '2024-01-20',
            'NÃO', '', 'SIM', '2024-02-01',
            'SIM', '2024-02-05', 'Procuração JUCEC', 'Todas válidas',
            
            # Bloco 8: Observações e Dados Adicionais
            'Cliente em dia com obrigações', '5', 'ATIVO',
            '2024-01-01', 'Sistema', 'ALTA',
            'VIP,PREMIUM', 'Cliente cadastrado via sistema'
        ]
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Escrever exemplo
        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        template_filename = f'template_importacao_clientes_{timestamp}.xlsx'
        
        from flask import send_file
        
        return send_file(
            buffer,
            as_attachment=True, 
            download_name=template_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erro ao gerar template: {e}")
        flash(f'Erro ao gerar template: {str(e)}', 'error')
        return redirect(url_for('import_page'))

def create_simple_template():
    """Cria template simples sem pandas como fallback"""
    try:
        import csv
        import io
        from flask import Response
        
        # Dados do template em formato CSV
        csv_data = [
            ['NOME DA EMPRESA', 'CT', 'FS', 'DP', 'COD. FORTES CT', 'COD. FORTES FS', 'COD. FORTES PS', 'COD. DOMÍNIO', 'SISTEMA UTILIZADO', 'MÓDULO SPED TRIER', 'RAZÃO SOCIAL NA RECEITA', 'NOME FANTASIA NA RECEITA', 'CNPJ', 'INSC. EST.', 'INSC. MUN.', 'SEGMENTO', 'ATIVIDADE', 'TRIBUTAÇÃO', 'PERFIL', 'CIDADE', 'DONO / RESP.', 'COD. ACESSO SIMPLES', 'CPF OU CNPJ', 'ACESSO ISS', 'ACESSO SEFIN', 'ACESSO SEUMA', 'ACESSO EMP. WEB', 'SENHA EMP. WEB', 'ACESSO FAP/INSS', 'ACESSO CRF', 'EMAIL GESTOR', 'ANVISA GESTOR', 'ANVISA EMPRESA', 'ACESSO IBAMA', 'ACESSO SEMACE', 'SENHA SEMACE', 'PROC. RC', 'PROC. CX', 'PROC. SW', 'MÊS/ANO DE  INÍCIO', 'RESPONSÁVEL IMEDIATO', 'TELEFONE FIXO', 'TELEFONE CELULAR', 'E-MAILS', 'SÓCIO', 'SÓCIO.1', 'SÓCIO.2', 'SÓCIO.3'],
            ['Exemplo Empresa Ltda', 'SIM', 'NAO', 'NAO', '123', '456', '789', '999', 'Sistema XYZ', 'Módulo ABC', 'EXEMPLO EMPRESA LTDA', 'Exemplo Empresa', '12.345.678/0001-99', '123456789', '987654321', 'COMÉRCIO', 'VAREJO', 'SIMPLES', 'A', 'FORTALEZA', 'João Silva', 'ABC123', '123.456.789-00', 'usuario123', 'senha123', 'acesso123', 'web123', 'senha456', 'fap123', 'crf123', 'gestor@empresa.com', 'anvisa123', 'empresa123', 'ibama123', 'semace123', 'senha789', 'OK', 'OK', 'OK', '01/2024', 'João Silva', '(85) 3333-4444', '(85) 99999-8888', 'contato@empresa.com', 'João Silva', 'Maria Silva', '', '']
        ]
        
        # Criar CSV em memória
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        
        # Converter para bytes
        csv_bytes = output.getvalue().encode('utf-8-sig')  # BOM para Excel
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'template_importacao_clientes_{timestamp}.csv'
        
        return Response(
            csv_bytes,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        print(f"❌ Erro ao criar template CSV: {e}")
        flash('Erro ao gerar template. Crie manualmente um arquivo Excel com as colunas necessárias.', 'error')
        return redirect(url_for('import_page'))

# Removido: Seção completa de ROTAS DE RELATÓRIOS (funcionalidade Power BI removida)

def calculate_dashboard_stats(clients):
    """Calcula estatísticas para o dashboard baseado nos dados SIGEC"""
    stats = {
        'total_clientes': len(clients),
        'clientes_ativos': 0,
        'empresas': 0,
        'domesticas': 0,
        'mei': 0,
        'simples_nacional': 0,
        'lucro_presumido': 0,
        'lucro_real': 0,
        'ct': 0,
        'fs': 0,
        'dp': 0,
        'bpo': 0
    }
    
    for client in clients:
        # Contadores básicos
        if client.get('ativo', True):
            stats['clientes_ativos'] += 1
        
        # Serviços
        if client.get('ct', False):
            stats['ct'] += 1
        if client.get('fs', False):
            stats['fs'] += 1
        if client.get('dp', False):
            stats['dp'] += 1
        if client.get('bpoFinanceiro', False):
            stats['bpo'] += 1
        
        # Categorização por regime federal (SIGEC)
        regime = str(client.get('regimeFederal', '')).upper().strip()
        perfil = str(client.get('perfil', '')).upper().strip()
        
        # MEI - Microempreendedor Individual
        if 'MEI' in regime or 'MICROEMPRESARIO' in perfil or 'INDIVIDUAL' in perfil:
            stats['mei'] += 1
        # Simples Nacional
        elif 'SIMPLES' in regime or 'SN' in regime:
            stats['simples_nacional'] += 1
        # Lucro Presumido
        elif 'PRESUMIDO' in regime or 'LP' in regime:
            stats['lucro_presumido'] += 1
        # Lucro Real
        elif 'REAL' in regime or 'LR' in regime:
            stats['lucro_real'] += 1
        # Domésticas - identificar por perfil ou atividade
        elif 'DOMESTICA' in perfil or 'EMPREGADA' in perfil:
            stats['domesticas'] += 1
        else:
            # Demais casos consideramos como empresas
            stats['empresas'] += 1
    
    return stats

def calculate_dashboard_stats_optimized(clients):
    """Versão otimizada para ambientes com pouca memória"""
    if not clients:
        return {
            'total_clientes': 0, 'clientes_ativos': 0, 'empresas': 0, 
            'domesticas': 0, 'mei': 0, 'simples_nacional': 0,
            'lucro_presumido': 0, 'lucro_real': 0,
            'ct': 0, 'fs': 0, 'dp': 0, 'bpo': 0
        }
    
    # Inicializar contadores
    stats = {
        'total_clientes': len(clients),
        'clientes_ativos': 0,
        'empresas': 0,
        'domesticas': 0,
        'mei': 0,
        'simples_nacional': 0,
        'lucro_presumido': 0,
        'lucro_real': 0,
        'ct': 0,
        'fs': 0,
        'dp': 0,
        'bpo': 0
    }
    
    # Processar em lotes para economizar memória
    batch_size = get_optimized_batch_size() if MEMORY_OPTIMIZER_AVAILABLE else 50
    
    for i in range(0, len(clients), batch_size):
        batch = clients[i:i+batch_size]
        
        for client in batch:
            # Contadores básicos (apenas campos essenciais)
            if client.get('ativo', True):
                stats['clientes_ativos'] += 1
            
            # Serviços básicos
            if client.get('ct'):
                stats['ct'] += 1
            if client.get('fs'):
                stats['fs'] += 1
            if client.get('dp'):
                stats['dp'] += 1
            if client.get('bpoFinanceiro'):
                stats['bpo'] += 1
            
            # Categorização simplificada (menos processamento de string)
            regime = client.get('regimeFederal', '')
            if regime:
                regime_upper = regime.upper()
                if 'MEI' in regime_upper:
                    stats['mei'] += 1
                elif 'SIMPLES' in regime_upper:
                    stats['simples_nacional'] += 1
                elif 'PRESUMIDO' in regime_upper:
                    stats['lucro_presumido'] += 1
                elif 'REAL' in regime_upper:
                    stats['lucro_real'] += 1
                else:
                    stats['empresas'] += 1
            else:
                stats['empresas'] += 1
        
        # Limpeza de memória após cada lote
        if os.environ.get('FLASK_ENV') == 'production':
            gc.collect()
    
    return stats

@app.route('/api/users')
@admin_required
def get_users():
    """API para obter lista de usuários disponíveis"""
    user_svc = get_user_service()
    if user_svc:
        users = user_svc.get_available_users()
        return jsonify({'users': users})
    else:
        return jsonify({'users': ['todos', 'admin', 'usuario']})

@app.route('/api/memory-status')
@admin_required  
def memory_status():
    """API para monitorar uso de memória em produção"""
    try:
        # Usar monitor ultra-otimizado se disponível
        try:
            from ultra_memory_monitor import get_memory_status
            memory_info = get_memory_status()
        except ImportError:
            # Fallback para método básico
            memory_info = {
                'timestamp': datetime.now().isoformat(),
                'environment': os.environ.get('FLASK_ENV', 'development'),
                'python_version': sys.version.split()[0],
                'monitoring': False
            }
            
            # Tentar obter info básica de memória
            try:
                import psutil
                process = psutil.Process()
                memory_info.update({
                    'memory_mb': round(process.memory_info().rss / 1024 / 1024, 1),
                    'memory_percent': round(process.memory_percent(), 1),
                    'cpu_percent': round(process.cpu_percent(), 1),
                    'threads': process.num_threads(),
                })
                
                # Alertas baseados nos limites do Render (512MB)
                memory_mb = memory_info['memory_mb']
                if memory_mb > 450:
                    memory_info['alert'] = 'CRITICAL - Próximo do limite de 512MB'
                    memory_info['alert_level'] = 'danger'
                elif memory_mb > 350:
                    memory_info['alert'] = 'WARNING - Uso de memória elevado'
                    memory_info['alert_level'] = 'warning'
                else:
                    memory_info['alert'] = 'OK - Uso de memória normal'
                    memory_info['alert_level'] = 'success'
                    
            except ImportError:
                memory_info['memory_mb'] = 'N/A (psutil não disponível)'
                memory_info['alert'] = 'Monitoramento limitado'
                memory_info['alert_level'] = 'info'
        
        # Informações sobre garbage collection
        memory_info['gc_counts'] = gc.get_count()
        memory_info['gc_threshold'] = gc.get_threshold()
        
        # Informações sobre lazy loading
        services_loaded = {
            'storage_service': storage_service is not None,
            'meeting_service': meeting_service is not None,
            'user_service': user_service is not None,
            # Removido: 'report_service': report_service is not None,
            'import_service': import_service is not None
        }
        memory_info['services_loaded'] = services_loaded
        memory_info['services_count'] = sum(services_loaded.values())
        
        # Configurações de otimização ativas
        optimizations = {
            'ultra_optimizer_available': MEMORY_OPTIMIZER_AVAILABLE,
            'max_content_length': app.config.get('MAX_CONTENT_LENGTH', 0) / 1024,  # KB
            'json_sort_keys': app.config.get('JSON_SORT_KEYS', True),
            'web_concurrency': os.environ.get('WEB_CONCURRENCY', 'auto'),
            'worker_connections': os.environ.get('WORKER_CONNECTIONS', 'auto'),
        }
        memory_info['optimizations'] = optimizations
        
        return jsonify(memory_info)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'timestamp': datetime.now().isoformat(),
            'alert': 'ERROR - Falha ao obter status de memória',
            'alert_level': 'danger'
        }), 500

@app.route('/api/force-cleanup')
@admin_required
def force_cleanup_api():
    """API para forçar limpeza de memória"""
    try:
        # Usar monitor ultra se disponível
        try:
            from ultra_memory_monitor import force_cleanup
            force_cleanup()
            message = "Limpeza ULTRA executada com sucesso"
        except ImportError:
            # Fallback para limpeza básica
            if MEMORY_OPTIMIZER_AVAILABLE:
                UltraMemoryOptimizer.force_cleanup()
                message = "Limpeza básica executada com sucesso"
            else:
                for _ in range(3):
                    gc.collect()
                message = "Limpeza manual executada"
        
        return jsonify({
            'success': True,
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/api/auth-status')
@admin_required
def auth_status():
    """API para diagnóstico de autenticação Google Sheets"""
    try:
        diagnosis = {
            'timestamp': datetime.now().isoformat(),
            'environment': os.environ.get('FLASK_ENV', 'development'),
            'google_sheets_config': {
                'use_google_sheets': USE_GOOGLE_SHEETS,
                'use_service_account': USE_SERVICE_ACCOUNT,
                'use_oauth2': USE_OAUTH2,
                'sheets_id_configured': bool(GOOGLE_SHEETS_ID),
                'api_key_configured': bool(GOOGLE_SHEETS_API_KEY),
            }
        }
        
        # Verificar variável de ambiente do Service Account
        service_account_json = os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON')
        diagnosis['service_account'] = {
            'env_var_present': bool(service_account_json),
            'env_var_length': len(service_account_json) if service_account_json else 0,
            'is_valid_json': False,
            'project_id': None,
            'client_email': None
        }
        
        if service_account_json:
            try:
                credentials_info = json.loads(service_account_json)
                diagnosis['service_account'].update({
                    'is_valid_json': True,
                    'project_id': credentials_info.get('project_id'),
                    'client_email': credentials_info.get('client_email'),
                    'has_private_key': bool(credentials_info.get('private_key'))
                })
            except json.JSONDecodeError as e:
                diagnosis['service_account']['json_error'] = str(e)
        
        # Verificar status do storage service
        storage = get_storage_service()
        diagnosis['storage_service'] = {
            'type': type(storage).__name__,
            'is_fallback': hasattr(storage, 'is_fallback') and storage.is_fallback,
            'initialized': storage is not None
        }
        
        # Se estiver usando fallback, incluir informações adicionais
        if hasattr(storage, 'get_status_info'):
            diagnosis['fallback_info'] = storage.get_status_info()
        
        return jsonify(diagnosis)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'timestamp': datetime.now().isoformat(),
            'status': 'ERROR - Falha no diagnóstico de autenticação'
        }), 500

# Rota raiz pública - redireciona conforme autenticação  
@app.route('/home')
def home():
    """Página inicial pública - redireciona baseado no status de autenticação"""
    print("🏠 HOME: Verificando status de autenticação...")
    
    # Se já está logado, redirecionar para dashboard principal
    if 'user_id' in session and session.get('user_id'):
        print(f"🏠 HOME: Usuário já autenticado (ID: {session['user_id']}) - Redirecionando para dashboard")
        return redirect(url_for('index'))
    
    # Se não está logado, redirecionar para login
    print("🏠 HOME: Usuário não autenticado - Redirecionando para login")
    return redirect(url_for('login'))

@app.route('/')
def index():
    """Rota raiz - verifica autenticação e redireciona conforme necessário"""
    print("🏠 INDEX: Verificando status de autenticação...")
    
    # Se não está logado, redirecionar para login
    if 'user_id' not in session or not session.get('user_id'):
        print("🏠 INDEX: Usuário não autenticado - Redirecionando para login")
        flash('Você precisa fazer login para acessar esta página.', 'warning')
        return redirect(url_for('login'))
    
    # Se está logado, continuar com a lógica original
    print(f"🏠 INDEX: Usuário autenticado (ID: {session['user_id']}) - Carregando dashboard")
    print("🔍 === ROTA INDEX CHAMADA (ULTRA-MEMORY OPTIMIZED) ===")
    
    # Obter filtro de status da URL (padrão: apenas ativos)
    status_filter = request.args.get('status', 'ativo')
    print(f"🔍 Filtro de status aplicado: {status_filter}")
    
    try:
        print("📊 Carregando clientes com lazy loading EXTREMO...")
        
        # OTIMIZAÇÃO MEMÓRIA: Forçar limpeza antes de carregar
        if os.environ.get('FLASK_ENV') == 'production':
            gc.collect()
            gc.collect()  # Dupla passada
        
        # OTIMIZAÇÃO MEMÓRIA: Usar lazy loading e limite ULTRA-restritivo
        storage = get_storage_service()
        
        # Verificar se é possível usar serviço otimizado
        try:
            from services.memory_optimized_sheets_service import MemoryOptimizedGoogleSheetsService
            if hasattr(storage, 'spreadsheet_id'):
                print("🧠 Usando serviço ULTRA-otimizado para memória")
                optimized_service = MemoryOptimizedGoogleSheetsService(
                    storage.spreadsheet_id, 
                    storage.range_name
                )
                clients = optimized_service.get_clients()
            else:
                clients = storage.get_clients()
        except ImportError:
            clients = storage.get_clients()
        
        # Aplicar filtro de status ANTES da limitação de memória
        original_count = len(clients)
        if status_filter == 'ativo':
            clients = [c for c in clients if c.get('ativo', True) and c.get('statusCliente', 'ativo').lower() == 'ativo']
            print(f"🔍 Filtro ATIVO aplicado: {len(clients)} de {original_count} clientes")
        elif status_filter == 'inativo':
            clients = [c for c in clients if not c.get('ativo', True) or c.get('statusCliente', 'ativo').lower() == 'inativo']
            print(f"🔍 Filtro INATIVO aplicado: {len(clients)} de {original_count} clientes")
        elif status_filter == 'todos':
            print(f"🔍 Filtro TODOS aplicado: {len(clients)} clientes (sem filtro)")
        
        # Limite EXTREMO baseado na memória disponível
        max_clients = ULTRA_MEMORY_SETTINGS.get('MAX_ROWS_PER_REQUEST', 10) if MEMORY_OPTIMIZER_AVAILABLE else 10
        
        # Para produção, ser EXTREMAMENTE restritivo
        if os.environ.get('FLASK_ENV') == 'production':
            max_clients = min(max_clients, 5)  # Máximo 5 clientes em produção
            
        if len(clients) > max_clients:
            clients = clients[:max_clients]  # Truncar para economizar memória extrema
            print(f"🧠 EXTREMAMENTE LIMITADO a {max_clients} clientes (economia RAM crítica)")
        
        print(f"✅ {len(clients)} clientes carregados")
        print(f"💾 Memória atual: {UltraMemoryOptimizer.get_memory_usage() if MEMORY_OPTIMIZER_AVAILABLE else 'N/A'}")
        
        # OTIMIZAÇÃO MEMÓRIA: Stats calculadas corretamente
        try:
            # Calcular estatísticas reais mantendo otimização de memória
            stats = calculate_dashboard_stats_optimized(clients)
            print(f"📈 Estatísticas calculadas: {stats['total_clientes']} total, {stats['mei']} MEI, {stats['simples_nacional']} SN, {stats['lucro_presumido']} LP, {stats['lucro_real']} LR")
        except Exception as stats_error:
            print(f"⚠️ Erro ao calcular stats: {stats_error}")
            stats = {
                'total_clientes': len(clients), 
                'clientes_ativos': len(clients),  # Simplificado
                'empresas': len(clients), 'domesticas': 0, 'mei': 0, 'simples_nacional': 0,
                'lucro_presumido': 0, 'lucro_real': 0,
                'ct': 0, 'fs': 0, 'dp': 0, 'bpo': 0
            }
        
        # Garbage collection AGRESSIVO após processamento
        if os.environ.get('FLASK_ENV') == 'production':
            # Múltiplas passadas para garantir limpeza máxima
            for _ in range(3):
                gc.collect()
            print(f"💾 Memória pós-GC-EXTREMO: {MemoryOptimizer.get_memory_usage() if MEMORY_OPTIMIZER_AVAILABLE else 'N/A'}")
        
        return render_template('index_modern.html', clients=clients, stats=stats, status_filter=status_filter)
        
    except Exception as e:
        print(f"❌ ERRO na rota index: {str(e)}")
        print(f"🔍 Tipo do erro: {type(e).__name__}")
        print(f"🔍 Status filter no except: {status_filter}")
        flash(f'Erro ao carregar clientes: {str(e)}', 'error')
        
        # Em caso de erro, criar stats vazias e forçar limpeza
        stats = {
            'total_clientes': 0, 'clientes_ativos': 0, 'empresas': 0,
            'domesticas': 0, 'mei': 0, 'simples_nacional': 0,
            'lucro_presumido': 0, 'lucro_real': 0,
            'ct': 0, 'fs': 0, 'dp': 0, 'bpo': 0
        }
        
        # Limpeza de emergência
        if os.environ.get('FLASK_ENV') == 'production':
            for _ in range(5):
                gc.collect()
        
        return render_template('index_modern.html', clients=[], stats=stats, status_filter=status_filter)

@app.route('/test')
@login_required
def test():
    """Rota de teste simples"""
    return """
    <html>
    <head>
        <title>Teste</title>
        <style>
            body { font-family: Arial; padding: 20px; background: #f0f0f0; }
            .container { background: white; padding: 20px; border-radius: 8px; }
            h1 { color: #333; }
            .success { color: green; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🚀 Aplicação Flask Funcionando!</h1>
            <p class="success">✅ Servidor está respondendo corretamente</p>
            <p>📊 Google Sheets configurado</p>
            <p>🔧 Sistema híbrido ativo</p>
            <a href="/">← Voltar para página principal</a>
        </div>
    </body>
    </html>
    """

@app.route('/update-sheet-headers')
@admin_required
def update_sheet_headers():
    """Atualiza os cabeçalhos da planilha Google Sheets removendo campos não utilizados"""
    try:
        print("🔧 [ADMIN] Iniciando atualização dos cabeçalhos da planilha...")
        
        # Usar get_storage_service() para lazy loading
        storage = get_storage_service()
        
        # Verificar se é o serviço do Google Sheets
        if hasattr(storage, 'update_sheet_headers_for_removed_fields'):
            print("✅ [ADMIN] Serviço Google Sheets detectado, iniciando atualização...")
            success = storage.update_sheet_headers_for_removed_fields()
            
            if success:
                message = """
                <html>
                <head>
                    <title>Atualização Concluída</title>
                    <style>
                        body { font-family: Arial; padding: 20px; background: #f0f0f0; }
                        .container { background: white; padding: 20px; border-radius: 8px; max-width: 800px; margin: 0 auto; }
                        h1 { color: #28a745; }
                        .info { background: #e7f3ff; padding: 15px; border-radius: 5px; margin: 10px 0; }
                        .removed-fields { background: #fff3cd; padding: 15px; border-radius: 5px; margin: 10px 0; }
                        ul { text-align: left; }
                        .back-btn { background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block; margin-top: 20px; }
                    </style>
                </head>
                <body>
                    <div class="container">
                        <h1>✅ Planilha Google Sheets Atualizada!</h1>
                        
                        <div class="info">
                            <h3>📊 Resumo da Atualização:</h3>
                            <p><strong>Colunas atuais:</strong> 86 (antes eram 97)</p>
                            <p><strong>Campos removidos:</strong> 11 campos não utilizados</p>
                            <p><strong>Status:</strong> Planilha otimizada e sincronizada com o sistema</p>
                        </div>
                        
                        <div class="removed-fields">
                            <h3>🗑️ Campos Removidos do Bloco 2:</h3>
                            <ul>
                                <li>Sistema Principal</li>
                                <li>Versão do Sistema</li>
                                <li>Código Acesso Simples</li>
                                <li>CPF/CNPJ para Acesso</li>
                                <li>Portal Cliente Ativo</li>
                                <li>Integração Domínio</li>
                                <li>Sistema Onvio</li>
                                <li>Onvio Contábil</li>
                                <li>Onvio Fiscal</li>
                                <li>Onvio Pessoal</li>
                                <li>Módulo SPED Trier</li>
                            </ul>
                        </div>
                        
                        <div class="info">
                            <h3>✅ Benefícios:</h3>
                            <ul>
                                <li>Planilha mais limpa e organizada</li>
                                <li>Melhor performance</li>
                                <li>Campos alinhados com o que realmente é usado</li>
                                <li>Dados existentes preservados</li>
                            </ul>
                        </div>
                        
                        <a href="/" class="back-btn">← Voltar ao Painel</a>
                    </div>
                </body>
                </html>
                """
                return message
            else:
                return """
                <html>
                <body style="font-family: Arial; padding: 20px; text-align: center;">
                    <h1 style="color: #dc3545;">❌ Erro na Atualização</h1>
                    <p>Não foi possível atualizar os cabeçalhos da planilha.</p>
                    <p>Verifique os logs do sistema para mais detalhes.</p>
                    <a href="/" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">← Voltar</a>
                </body>
                </html>
                """
        else:
            return """
            <html>
            <body style="font-family: Arial; padding: 20px; text-align: center;">
                <h1 style="color: #ffc107;">⚠️ Serviço Não Compatível</h1>
                <p>Esta função só está disponível quando usando Google Sheets como storage.</p>
                <p>Serviço atual: {}</p>
                <a href="/" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">← Voltar</a>
            </body>
            </html>
            """.format(type(storage).__name__)
            
    except Exception as e:
        print(f"❌ [ADMIN] Erro ao atualizar cabeçalhos: {e}")
        return f"""
        <html>
        <body style="font-family: Arial; padding: 20px; text-align: center;">
            <h1 style="color: #dc3545;">❌ Erro</h1>
            <p>Erro ao atualizar cabeçalhos: {str(e)}</p>
            <a href="/" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">← Voltar</a>
        </body>
        </html>
        """

@app.route('/client/<client_id>')
@login_required
def view_client(client_id):
    try:
        print(f"🔍 [VIEW] ===== CARREGANDO CLIENTE PARA VISUALIZAÇÃO =====")
        print(f"🔍 [VIEW] ID solicitado: '{client_id}'")
        print(f"🔍 [VIEW] Tipo do ID: {type(client_id)}")
        
        # CORREÇÃO: Usar get_storage_service() para lazy loading
        storage = get_storage_service()
        client = storage.get_client(client_id)
        print(f"🔍 [VIEW] Cliente carregado: {client is not None}")
        
        if client:
            print(f"🔍 [VIEW] Nome do cliente: {client.get('nomeEmpresa')}")
            print(f"🔍 [VIEW] ID do cliente retornado: '{client.get('id')}'")
            
            # DEBUG: Verificar campos de sócio disponíveis
            print("🔍 [VIEW] ===== DEBUG SÓCIOS =====")
            for key in client.keys():
                if 'socio' in key.lower():
                    print(f"🔍 [VIEW] Campo sócio: {key} = {client[key]}")
            print("🔍 [VIEW] ========================")
            
            # DEBUG: Verificar TODOS os campos do cliente
            print("🔍 [VIEW] ===== DEBUG TODOS OS CAMPOS =====")
            print(f"🔍 [VIEW] Total de campos: {len(client.keys())}")
            for key in sorted(client.keys()):
                value = client[key]
                if len(str(value)) > 50:
                    value_display = str(value)[:47] + "..."
                else:
                    value_display = value
                print(f"🔍 [VIEW] {key}: {value_display}")
            print("🔍 [VIEW] ===================================")
            
            # DEBUG ESPECÍFICO: Testar os diferentes formatos de campo de sócio
            print("🔍 [VIEW] ===== TESTE ESPECÍFICO SÓCIOS =====")
            test_patterns = [
                'socio_1_nome', 'socio1_nome', 'socio1',
                'socio_1_cpf', 'socio1_cpf', 'socio1_Cpf',
                'SÓCIO 1 NOME', 'sócio_1_nome'
            ]
            for pattern in test_patterns:
                if pattern in client:
                    print(f"🔍 [VIEW] ENCONTRADO: {pattern} = {client[pattern]}")
                else:
                    print(f"🔍 [VIEW] NÃO EXISTE: {pattern}")
            print("🔍 [VIEW] ========================================")
            
            # CORREÇÃO: Usar client_view_modern_new.html como template principal
            return render_template('client_view_modern_new.html', client=client)
        else:
            print(f"❌ [VIEW] Cliente {client_id} não encontrado!")
            flash('Cliente não encontrado', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        print(f"❌ [VIEW] Erro ao carregar cliente: {str(e)}")
        import traceback
        print(f"❌ [VIEW] Traceback: {traceback.format_exc()}")
        flash(f'Erro ao carregar cliente: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/client/new')
@login_required
def new_client():
    segmentos = get_segmentos_list()
    atividades = get_atividades_list()
    sistemas = get_sistemas_list()
    return render_template('client_form_complete.html', client=None, segmentos=segmentos, atividades=atividades, sistemas=sistemas)

@app.route('/client/<client_id>/edit')
@login_required
def edit_client(client_id):
    try:
        print(f"🔍 [EDIT] ===== CARREGANDO CLIENTE PARA EDIÇÃO =====")
        print(f"🔍 [EDIT] ID solicitado: '{client_id}'")
        
        # CORREÇÃO: Usar get_storage_service() para lazy loading
        storage = get_storage_service()
        client = storage.get_client(client_id)
        print(f"🔍 [EDIT] Cliente carregado: {client is not None}")
        
        if client:
            print(f"🔍 [EDIT] Nome do cliente: {client.get('nomeEmpresa')}")
            print(f"🔍 [EDIT] ID do cliente retornado: '{client.get('id')}'")
            print(f"🔍 [EDIT] Tipo do ID: {type(client.get('id'))}")
            print(f"🔍 [EDIT] Dados principais: {list(client.keys())[:10]}")
            
            # Garantir que o ID está correto
            if not client.get('id'):
                print(f"⚠️ [EDIT] Cliente não tem ID! Forçando ID = {client_id}")
                client['id'] = client_id
            
            segmentos = get_segmentos_list()
            atividades = get_atividades_list()
            sistemas = get_sistemas_list()
            return render_template('client_form_complete.html', client=client, segmentos=segmentos, atividades=atividades, sistemas=sistemas)
        else:
            print(f"❌ [EDIT] Cliente {client_id} não encontrado!")
            flash('Cliente não encontrado', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        print(f"❌ [EDIT] Erro ao carregar cliente: {str(e)}")
        import traceback
        print(f"❌ [EDIT] Traceback: {traceback.format_exc()}")
        flash(f'Erro ao carregar cliente: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/client', methods=['POST'])
@login_required
def save_client():
    import re  # CORREÇÃO 03: Import necessário para normalização CPF/CNPJ
    
    print("🔍 === FUNÇÃO SAVE_CLIENT CHAMADA ===")
    print(f"🔍 Método da requisição: {request.method}")
    print(f"🔍 Dados do form: {dict(request.form)}")
    
    # CORREÇÃO DUPLICAÇÃO: Verificar ID primeiro
    client_id = request.form.get('id', '').strip()
    row_number = request.form.get('row_number')
    print(f"🔍 ID do cliente (raw): '{request.form.get('id')}'")
    print(f"🔍 ID do cliente (processed): '{client_id}'")
    print(f"🔍 Operação: {'EDIÇÃO' if client_id else 'CRIAÇÃO'}")
    
    # DEBUG ESPECÍFICO PARA CAMPOS DE CÓDIGOS
    print("🔍 === VERIFICANDO CAMPOS DE CÓDIGOS ===")
    print(f"🔍 codDominio (raw): '{request.form.get('codDominio')}'")
    print(f"🔍 codFortesCt (raw): '{request.form.get('codFortesCt')}'") 
    print(f"🔍 codFortesFs (raw): '{request.form.get('codFortesFs')}'")
    print(f"🔍 codFortesPs (raw): '{request.form.get('codFortesPs')}'")
    print("🔍 =======================================")
    
    try:
        # Validar dados obrigatórios do Bloco 1
        nome_empresa = request.form.get('nomeEmpresa', '').strip()
        razao_social = request.form.get('razaoSocialReceita', '').strip()
        nome_fantasia = request.form.get('nomeFantasiaReceita', '').strip()
        cpf_cnpj = request.form.get('cpfCnpj', '').strip()
        # Compat: o formulário usa 'perfilCliente'
        perfil = (request.form.get('perfilCliente') or request.form.get('perfil') or '').strip()
        insc_est = request.form.get('inscEst', '').strip()
        insc_mun = request.form.get('inscMun', '').strip()
        estado = request.form.get('estado', '').strip()
        cidade = request.form.get('cidade', '').strip()
        regime_federal = request.form.get('regimeFederal', '').strip()
        regime_estadual = request.form.get('regimeEstadual', '').strip()
        segmento = request.form.get('segmento', '').strip()
        atividade = request.form.get('atividade', '').strip()
        
        # DEBUG ESPECÍFICO: Campos que não estão salvando
        print("🔍 === DEBUG CAMPOS ESPECÍFICOS ===")
        print(f"🔍 bpoFinanceiro (form): '{request.form.get('bpoFinanceiro')}'")
        print(f"🔍 ct (form): '{request.form.get('ct')}'")
        print(f"🔍 fs (form): '{request.form.get('fs')}'")
        print(f"🔍 dp (form): '{request.form.get('dp')}'")
        print(f"🔍 codDominio (form): '{request.form.get('codDominio')}'")
        print(f"🔍 codFortesCt (form): '{request.form.get('codFortesCt')}'")
        print(f"🔍 codFortesFs (form): '{request.form.get('codFortesFs')}'")
        print(f"🔍 codFortesPs (form): '{request.form.get('codFortesPs')}'")
        print("🔍 ====================================")
        
        # Verificar se há outras chaves no formulário
        
        # Função auxiliar para retornar ao formulário com dados preservados
        def return_to_form_with_error(error_msg):
            flash(error_msg, 'error')
            # Construir dados do cliente a partir do formulário para preservar
            form_data = {}
            for key in request.form:
                form_data[key] = request.form[key]
            
            # Se tem ID, é edição
            if client_id:
                form_data['id'] = client_id
            
            segmentos = get_segmentos_list()
            atividades = get_atividades_list()
            sistemas = get_sistemas_list()
            return render_template('client_form_complete.html', 
                                 client=form_data, 
                                 segmentos=segmentos, 
                                 atividades=atividades,
                                 sistemas=sistemas)
        
        # Validações obrigatórias - retornar ao formulário em caso de erro
        if not nome_empresa:
            return return_to_form_with_error('❌ Campo obrigatório: Nome da empresa deve ser preenchido.')
        if not razao_social:
            return return_to_form_with_error('❌ Campo obrigatório: Razão Social (Receita) deve ser preenchida.')
        if not nome_fantasia:
            return return_to_form_with_error('❌ Campo obrigatório: Nome Fantasia (Receita) deve ser preenchido.')
        if not cpf_cnpj:
            return return_to_form_with_error('❌ Campo obrigatório: CPF/CNPJ deve ser preenchido.')
        if not perfil:
            return return_to_form_with_error('❌ Campo obrigatório: Perfil do cliente deve ser selecionado.')
        if not insc_est:
            return return_to_form_with_error('❌ Campo obrigatório: Inscrição Estadual deve ser preenchida (ou "ISENTO").')
        if not insc_mun:
            return return_to_form_with_error('❌ Campo obrigatório: Inscrição Municipal deve ser preenchida (ou "ISENTO").')
        if not estado:
            return return_to_form_with_error('❌ Campo obrigatório: Estado deve ser selecionado.')
        if not cidade:
            return return_to_form_with_error('❌ Campo obrigatório: Cidade deve ser preenchida.')
        if not regime_federal:
            return return_to_form_with_error('❌ Campo obrigatório: Regime Federal deve ser selecionado.')
        if not regime_estadual:
            return return_to_form_with_error('❌ Campo obrigatório: Regime Estadual deve ser selecionado.')
        if not segmento:
            return return_to_form_with_error('❌ Campo obrigatório: Segmento deve ser selecionado.')
        if not atividade:
            return return_to_form_with_error('❌ Campo obrigatório: Atividade Principal deve ser selecionada.')
        
        # CORREÇÃO 02: Converter campos específicos para MAIÚSCULAS
        nome_empresa = nome_empresa.upper()
        razao_social = razao_social.upper()
        nome_fantasia = nome_fantasia.upper()
        cidade = cidade.upper()
        insc_est = insc_est.upper()
        insc_mun = insc_mun.upper()
        
        # CORREÇÃO 03: Normalizar CPF/CNPJ preservando zeros à esquerda
        cpf_cnpj_digits = re.sub(r'\D', '', cpf_cnpj)  # Remove formatação
        if len(cpf_cnpj_digits) == 11:
            # CPF - garantir 11 dígitos com zeros à esquerda
            cpf_cnpj_normalized = cpf_cnpj_digits.zfill(11)
            # Reformatar: 12345678900 -> 123.456.789-00
            cpf_cnpj = f"{cpf_cnpj_normalized[:3]}.{cpf_cnpj_normalized[3:6]}.{cpf_cnpj_normalized[6:9]}-{cpf_cnpj_normalized[9:11]}"
            print(f"🔍 CPF normalizado com zeros à esquerda: {cpf_cnpj}")
        elif len(cpf_cnpj_digits) == 14:
            # CNPJ - garantir 14 dígitos com zeros à esquerda
            cpf_cnpj_normalized = cpf_cnpj_digits.zfill(14)
            # Reformatar: 12345678000190 -> 12.345.678/0001-90
            cpf_cnpj = f"{cpf_cnpj_normalized[:2]}.{cpf_cnpj_normalized[2:5]}.{cpf_cnpj_normalized[5:8]}/{cpf_cnpj_normalized[8:12]}-{cpf_cnpj_normalized[12:14]}"
            print(f"🔍 CNPJ normalizado: {cpf_cnpj}")
        else:
            print(f"🔍 CPF/CNPJ mantido como recebido (tamanho: {len(cpf_cnpj_digits)}): {cpf_cnpj}")
        
        print(f"🔍 Campos convertidos para maiúsculas - Nome: {nome_empresa}, Cidade: {cidade}")
        
        print(f"🔍 Nome da empresa: {nome_empresa}")
        
        # CORREÇÃO DUPLICAÇÃO: Garantir que o ID seja passado corretamente
        # Dados básicos obrigatórios - Bloco 1
        client_data = {
            'id': client_id if client_id else None,  # FIXADO: usar variável processada
            '_row_number': int(row_number) if row_number and row_number.isdigit() else None,
            
            # Bloco 1: Informações da Pessoa Física / Jurídica
            'nomeEmpresa': nome_empresa,
            'razaoSocialReceita': razao_social,
            'nomeFantasiaReceita': nome_fantasia,
            'cpfCnpj': cpf_cnpj,
            'cnpj': cpf_cnpj,  # Manter compatibilidade com campo antigo
            'perfil': perfil,
            'perfilCliente': perfil,  # alias
            'inscEst': insc_est,
            'inscMun': insc_mun,
            'estado': estado,
            'cidade': cidade,
            'regimeFederal': regime_federal,
            'tributacao': regime_federal,  # Manter compatibilidade com campo antigo
            'regimeEstadual': regime_estadual,
            'segmento': segmento,
            'atividade': atividade,
            
            # Bloco 2: Serviços Prestados pela Control
            'bpoFinanceiro': request.form.get('bpoFinanceiro') == 'on',
            'ct': request.form.get('ct') == 'on',
            'fs': request.form.get('fs') == 'on',
            'dp': request.form.get('dp') == 'on',
            'domestica': request.form.get('domestica', '').strip(),
            'dataInicioServicos': request.form.get('dataInicioServicos', '').strip(),
            'geraArquivoSped': request.form.get('geraArquivoSped', '').strip(),
            'sistemaUtilizado': request.form.get('sistemaUtilizado', '').strip(),
            
            # Códigos dos Sistemas (Bloco 2)
            'codigoDominio': request.form.get('codDominio', '').strip(),
            'codigoFortesCT': request.form.get('codFortesCt', '').strip(),
            'codigoFortesFS': request.form.get('codFortesFs', '').strip(),
            'codigoFortesPS': request.form.get('codFortesPs', '').strip(),
            
            # Bloco 4: Contatos
            'telefoneFixo': request.form.get('telefoneFixo', ''),
            'telefoneCelular': request.form.get('telefoneCelular', ''),
            'whatsapp': request.form.get('whatsapp', ''),
            'emailPrincipal': request.form.get('emailPrincipal', ''),
            'emailSecundario': request.form.get('emailSecundario', ''),
            'responsavelImediato': request.form.get('responsavelImediato', ''),
            'emailsSocios': request.form.get('emailsSocios', ''),
            'contatoContador': request.form.get('contatoContador', ''),
            'telefoneContador': request.form.get('telefoneContador', ''),
            'emailContador': request.form.get('emailContador', ''),
        }
        
        # DEBUG ESPECÍFICO: Verificar valores processados
        print("🔍 === DEBUG VALORES PROCESSADOS ===")
        print(f"🔍 bpoFinanceiro (processado): {client_data.get('bpoFinanceiro')}")
        print(f"🔍 ct (processado): {client_data.get('ct')}")
        print(f"🔍 fs (processado): {client_data.get('fs')}")
        print(f"🔍 dp (processado): {client_data.get('dp')}")
        print(f"🔍 codigoDominio (processado): '{client_data.get('codigoDominio')}'")
        print(f"🔍 codigoFortesCT (processado): '{client_data.get('codigoFortesCT')}'")
        print(f"🔍 codigoFortesFS (processado): '{client_data.get('codigoFortesFS')}'")
        print(f"🔍 codigoFortesPS (processado): '{client_data.get('codigoFortesPS')}'")
        print("🔍 =====================================")
        
        # Processar dados dos sócios dinamicamente
        print("🔍 Processando dados dos sócios...")
        for i in range(1, 11):  # Suporte para até 10 sócios
            nome_socio = request.form.get(f'socio_{i}_nome', '').strip()
            if nome_socio:  # Se há nome, processar os dados do sócio
                # CORREÇÃO 02: Converter nome do sócio para MAIÚSCULAS
                nome_socio = nome_socio.upper()
                
                client_data[f'socio_{i}_nome'] = nome_socio
                client_data[f'socio_{i}_cpf'] = request.form.get(f'socio_{i}_cpf', '').strip()
                client_data[f'socio_{i}_data_nascimento'] = request.form.get(f'socio_{i}_data_nascimento', '').strip()
                client_data[f'socio_{i}_participacao'] = request.form.get(f'socio_{i}_participacao', '').strip()
                client_data[f'socio_{i}_administrador'] = request.form.get(f'socio_{i}_administrador') in ['1', 'on']
                client_data[f'socio_{i}_resp_legal'] = request.form.get('representante_legal') == f'socio_{i}'
                client_data[f'socio_{i}_email'] = request.form.get(f'socio_{i}_email', '').strip()
                client_data[f'socio_{i}_telefone'] = request.form.get(f'socio_{i}_telefone', '').strip()
                
                # COMPATIBILIDADE: Adicionar também campos sem underscore para templates antigos
                client_data[f'socio{i}_nome'] = nome_socio
                client_data[f'socio{i}_cpf'] = client_data[f'socio_{i}_cpf']
                client_data[f'socio{i}_administrador'] = client_data[f'socio_{i}_administrador']
                client_data[f'socio{i}'] = nome_socio  # Para templates mais antigos
                
                print(f"🔍 Sócio {i}: {nome_socio} - CPF: {client_data[f'socio_{i}_cpf']} - Admin: {client_data[f'socio_{i}_administrador']}")
                print(f"🔍 Compatibilidade: socio{i}_nome = {client_data[f'socio{i}_nome']}")
        
        # Processar dados dos contatos dinamicamente
        print("🔍 Processando dados dos contatos...")
        for i in range(1, 11):  # Suporte para até 10 contatos
            nome_contato = request.form.get(f'contato_{i}_nome', '').strip()
            telefone_contato = request.form.get(f'contato_{i}_telefone', '').strip()
            email_contato = request.form.get(f'contato_{i}_email', '').strip()
            cargo_contato = request.form.get(f'contato_{i}_cargo', '').strip()
            
            if nome_contato or telefone_contato or email_contato:  # Se há pelo menos um dado, processar o contato
                client_data[f'contato_{i}_nome'] = nome_contato
                client_data[f'contato_{i}_telefone'] = telefone_contato
                client_data[f'contato_{i}_email'] = email_contato
                client_data[f'contato_{i}_cargo'] = cargo_contato
                print(f"🔍 Contato {i}: {nome_contato} - Cargo: {cargo_contato} - Tel: {telefone_contato} - Email: {email_contato}")
        
        # Debug específico para dados de contatos básicos
        print("🔍 === DEBUG DADOS DE CONTATOS BÁSICOS ===")
        contatos_basicos = {
            'telefoneFixo': request.form.get('telefoneFixo', ''),
            'telefoneCelular': request.form.get('telefoneCelular', ''),
            'whatsapp': request.form.get('whatsapp', ''),
            'emailPrincipal': request.form.get('emailPrincipal', ''),
            'emailSecundario': request.form.get('emailSecundario', ''),
            'responsavelImediato': request.form.get('responsavelImediato', ''),
            'emailsSocios': request.form.get('emailsSocios', ''),
            'contatoContador': request.form.get('contatoContador', ''),
            'telefoneContador': request.form.get('telefoneContador', ''),
            'emailContador': request.form.get('emailContador', ''),
        }
        for key, value in contatos_basicos.items():
            print(f"🔍 {key}: '{value}'")
        
        # Debug específico para senhas e credenciais
        print("🔍 === DEBUG SENHAS E CREDENCIAIS ===")
        senhas_credenciais = {
            'cnpjAcessoSn': request.form.get('cnpjAcessoSn', ''),
            'cpfRepLegal': request.form.get('cpfRepLegal', ''),
            'codigoAcessoSn': request.form.get('codigoAcessoSn', ''),
            'senhaIss': request.form.get('senhaIss', ''),
            'senhaSefin': request.form.get('senhaSefin', ''),
            'senhaSeuma': request.form.get('senhaSeuma', ''),
            'acessoEmpWeb': request.form.get('acessoEmpWeb', ''),
            'senhaEmpWeb': request.form.get('senhaEmpWeb', ''),
            'anvisaEmpresa': request.form.get('anvisaEmpresa', ''),
            'senhaAnvisaEmpresa': request.form.get('senhaAnvisaEmpresa', ''),
            'anvisaGestor': request.form.get('anvisaGestor', ''),
            'senhaAnvisaGestor': request.form.get('senhaAnvisaGestor', ''),
            'acessoCrf': request.form.get('acessoCrf', ''),
            'senhaFapInss': request.form.get('senhaFapInss', ''),
        }
        for key, value in senhas_credenciais.items():
            if value:  # Só mostrar se tiver valor
                print(f"🔍 {key}: '{value[:3]}...' (tamanho: {len(value)})")
            else:
                print(f"🔍 {key}: VAZIO")
        
        # Debug específico para procurações
        print("🔍 === DEBUG PROCURAÇÕES ===")
        procuracoes_debug = {
            'procReceita': request.form.get('procReceita'),
            'dataProcReceita': request.form.get('dataProcReceita', ''),
            'procDte': request.form.get('procDte'),
            'dataProcDte': request.form.get('dataProcDte', ''),
            'procCaixa': request.form.get('procCaixa'),
            'dataProcCaixa': request.form.get('dataProcCaixa', ''),
            'procEmpWeb': request.form.get('procEmpWeb'),
            'dataProcEmpWeb': request.form.get('dataProcEmpWeb', ''),
            'procDet': request.form.get('procDet'),
            'dataProcDet': request.form.get('dataProcDet', ''),
            'outrasProc': request.form.get('outrasProc', ''),
            'obsProcuracoes': request.form.get('obsProcuracoes', ''),
        }
        for key, value in procuracoes_debug.items():
            if 'proc' in key.lower() and key != 'outrasProc' and key != 'obsProcuracoes':
                # Para checkboxes, mostrar se chegou como 'on' ou None
                converted = value == 'on' if value else False
                print(f"🔍 {key}: '{value}' -> {converted}")
            else:
                print(f"🔍 {key}: '{value}'")
        
        # Continuar com outros dados
        client_data.update({
            
            # Bloco 5: Senhas e Credenciais (ORGANIZADO CONFORME TEMPLATE)
            # Linha 1: CNPJ Acesso Simples Nacional, CPF do Representante Legal, Código de Acesso Simples Nacional, Senha ISS
            'cnpjAcessoSn': request.form.get('cnpjAcessoSn', ''),
            'cpfRepLegal': request.form.get('cpfRepLegal', ''),
            'codigoAcessoSn': request.form.get('codigoAcessoSn', ''),
            'senhaIss': request.form.get('senhaIss', ''),
            
            # Linha 2: Senha SEFIN, Senha SEUMA, Acesso EmpWeb, Senha EmpWeb
            'senhaSefin': request.form.get('senhaSefin', ''),
            'senhaSeuma': request.form.get('senhaSeuma', ''),
            'acessoEmpWeb': request.form.get('acessoEmpWeb', ''),
            'senhaEmpWeb': request.form.get('senhaEmpWeb', ''),
            
            # Linha 3: Login ANVISA Empresa, Senha ANVISA Empresa, Login ANVISA Gestor, Senha ANVISA Gestor
            'anvisaEmpresa': request.form.get('anvisaEmpresa', ''),
            'senhaAnvisaEmpresa': request.form.get('senhaAnvisaEmpresa', ''),
            'anvisaGestor': request.form.get('anvisaGestor', ''),
            'senhaAnvisaGestor': request.form.get('senhaAnvisaGestor', ''),
            
            # Linha 4: Acesso CRF, Senha FAP/INSS
            'acessoCrf': request.form.get('acessoCrf', ''),
            'senhaFapInss': request.form.get('senhaFapInss', ''),
            
            # Bloco 6: Procurações (CORRIGIDO - alinhado com formulário)
            'procReceita': request.form.get('procReceita') == 'on',
            'dataProcReceita': request.form.get('dataProcReceita', ''),
            'procDte': request.form.get('procDte') == 'on',
            'dataProcDte': request.form.get('dataProcDte', ''),
            'procCaixa': request.form.get('procCaixa') == 'on',
            'dataProcCaixa': request.form.get('dataProcCaixa', ''),
            'procEmpWeb': request.form.get('procEmpWeb') == 'on',
            'dataProcEmpWeb': request.form.get('dataProcEmpWeb', ''),
            'procDet': request.form.get('procDet') == 'on',
            'dataProcDet': request.form.get('dataProcDet', ''),
            'outrasProc': request.form.get('outrasProc', ''),
            'obsProcuracoes': request.form.get('obsProcuracoes', ''),
            
            # Bloco 7: Observações e Dados Adicionais (apenas campos mantidos)
            'observacoes': request.form.get('observacoes', ''),
        })
        
        # CORREÇÃO: Para statusCliente, usar valor do formulário se fornecido
        if client_id:
            # Estamos editando - usar status do formulário se fornecido, senão preservar atual
            try:
                storage = get_storage_service()
                current_client = storage.get_client(client_id) if storage else None
                current_status = current_client.get('statusCliente', 'ativo') if current_client else 'ativo'
                
                # LÓGICA CORRETA: Priorizar valor do formulário
                form_status = request.form.get('statusCliente')
                if form_status:
                    client_data['statusCliente'] = form_status
                    print(f"🔍 EDIÇÃO - Status do formulário usado: '{form_status}'")
                else:
                    client_data['statusCliente'] = current_status
                    print(f"🔍 EDIÇÃO - Status atual preservado: '{current_status}'")
                    
            except Exception as e:
                print(f"❌ Erro ao buscar status atual: {e}")
                client_data['statusCliente'] = request.form.get('statusCliente', 'ativo')
        else:
            # Cliente novo - padrão ativo
            client_data['statusCliente'] = request.form.get('statusCliente', 'ativo')
            print(f"🔍 NOVO CLIENTE - Status padrão: {client_data['statusCliente']}")
        
        # Finalizar dados básicos
        data_inicio_value = request.form.get('dataInicioServicos', '')
        
        # CORREÇÃO 05: Validar e normalizar formato MM/AAAA
        if data_inicio_value:
            import re
            # Se contém apenas dígitos, aplicar formato MM/AAAA
            digits_only = ''.join(filter(str.isdigit, data_inicio_value))
            if len(digits_only) == 6:  # MMAAAA
                month = digits_only[:2]
                year = digits_only[2:]
                if 1 <= int(month) <= 12:
                    data_inicio_value = f"{month}/{year}"
                    print(f"🔍 [CORREÇÃO 05] Formato normalizado: '{data_inicio_value}'")
            elif re.match(r'^(0[1-9]|1[0-2])\/\d{4}$', data_inicio_value):
                print(f"🔍 [CORREÇÃO 05] Formato já correto: '{data_inicio_value}'")
            else:
                print(f"⚠️ [CORREÇÃO 05] Formato inválido: '{data_inicio_value}'")
        
        print(f"🔍 [CORREÇÃO 05] dataInicioServicos final: '{data_inicio_value}'")
        
        client_data.update({
            'ultimaAtualizacao': datetime.now().isoformat(),
            
            # Campos de compatibilidade (manter existentes)
            'dataInicioServicos': data_inicio_value,
            'mesAnoInicio': data_inicio_value,
        })
        
        # Regras complementares
        # Sincronizar statusCliente com ativo para compatibilidade (automático)
        status_cliente = client_data.get('statusCliente', 'ativo')
        client_data['ativo'] = status_cliente == 'ativo'
        
        # Aplicar regra Doméstica no backend (segurança): só permitido quando CPF completo (11 dígitos)
        import re
        digits = re.sub(r'\D', '', client_data.get('cpfCnpj', ''))
        if len(digits) != 11:
            client_data['domestica'] = 'NÃO'
            print(f"🔍 Doméstica forçada para NÃO - documento tem {len(digits)} dígitos (≠11)")
        else:
            print(f"🔍 Doméstica permitida - CPF válido com {len(digits)} dígitos")
        
        # CORREÇÃO DUPLICAÇÃO: Melhor controle de criação vs edição
        # Usar client_id do formulário para determinar operação, não o ID gerado automaticamente
        if not client_id or client_id == '':
            print("🔍 NOVO CLIENTE: Não incluir ID nos dados para forçar criação")
            client_data['criadoEm'] = datetime.now().isoformat()
            # IMPORTANTE: NÃO incluir ID nos dados para novo cliente - deixar o serviço gerar
            if 'id' in client_data:
                del client_data['id']
        else:
            print(f"🔍 EDITANDO CLIENTE: ID = {client_id}")
            client_data['id'] = client_id  # Usar o ID do formulário
            # Para edição, sempre manter o ultimaAtualizacao
            client_data['ultimaAtualizacao'] = datetime.now().isoformat()
        
        print(f"🔍 Cliente preparado: {client_data.get('nomeEmpresa')}")
        print(f"🔍 ID final do cliente: {client_data.get('id')}")
        print(f"🔍 Tipo de operação: {'EDIÇÃO' if client_data.get('id') else 'CRIAÇÃO'}")
        print("🔍 Verificando conexão com storage_service...")
        
        # CORREÇÃO: Usar get_storage_service() para lazy loading
        storage = get_storage_service()
        if not storage:
            print("❌ storage_service não está disponível!")
            return return_to_form_with_error('Erro: Serviço de armazenamento não disponível')
        
        print("🔍 Chamando storage_service.save_client...")
        print(f"🔍 client_data['id']: '{client_data.get('id')}'")
        print(f"🔍 client_data['nomeEmpresa']: '{client_data.get('nomeEmpresa')}'")
        print(f"🔍 Dados essenciais: ID={client_data.get('id')}, Nome={client_data.get('nomeEmpresa')}")
        
        success = storage.save_client(client_data)
        
        print(f"🔍 Resultado do salvamento: {success}")
        
        if success:
            if client_data.get('id'):
                flash('Cliente atualizado com sucesso!', 'success')
                print("✅ Flash message de atualização adicionada")
                # Redirecionar para a página de visualização do cliente para mostrar os dados atualizados
                return redirect(url_for('view_client', client_id=client_data.get('id')))
            else:
                flash('Cliente criado com sucesso!', 'success')
                print("✅ Flash message de criação adicionada")
                # Para novo cliente, ir para página inicial já está bom
                return redirect(url_for('index'))
        else:
            # Em caso de erro no salvamento, retornar ao formulário com dados preservados
            return return_to_form_with_error('Erro ao salvar cliente')
            
    except Exception as e:
        print(f"❌ EXCEÇÃO na função save_client: {str(e)}")
        print(f"❌ Tipo da exceção: {type(e).__name__}")
        import traceback
        print(f"❌ Traceback completo: {traceback.format_exc()}")
        
        # Em caso de exceção, retornar ao formulário com dados preservados
        return return_to_form_with_error(f'Erro ao salvar cliente: {str(e)}')

@app.route('/client/<client_id>/delete', methods=['POST'])
@login_required
def delete_client(client_id):
    try:
        # Verificar se o usuário é administrador
        if not session.get('user_perfil') or session.get('user_perfil').lower() != 'administrador':
            flash('Acesso negado. Apenas administradores podem excluir clientes.', 'danger')
            return redirect(url_for('view_client', client_id=client_id))
        
        # CORREÇÃO: Usar get_storage_service() para lazy loading
        storage = get_storage_service()
        # Suportar deleção direta por linha quando disponível
        row_number = request.form.get('row_number')
        if row_number and hasattr(storage, 'delete_client_by_row'):
            try:
                row_number_int = int(row_number)
            except ValueError:
                row_number_int = None
            if row_number_int and row_number_int > 1:
                success = storage.delete_client_by_row(row_number_int)
            else:
                success = storage.delete_client(client_id)
        else:
            success = storage.delete_client(client_id)
        if success:
            flash('Cliente excluído com sucesso!', 'success')
        else:
            flash('Erro ao excluir cliente', 'error')
    except Exception as e:
        flash(f'Erro ao excluir cliente: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/client/<client_id>/meeting')
@login_required
def register_meeting(client_id):
    """Página para registrar ata de reunião com cliente"""
    try:
        client = storage_service.get_client(client_id)
        if client:
            return render_template('meeting_form.html', client=client)
        else:
            flash('Cliente não encontrado', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'Erro ao carregar cliente: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/client/<client_id>/meeting', methods=['POST'])
@login_required
def save_meeting(client_id):
    """Salvar ata de reunião"""
    try:
        # Busca o nome do cliente
        client = storage_service.get_client(client_id)
        client_name = client.get('nomeFantasiaReceita') or client.get('nomeEmpresa') if client else 'Cliente'
        
        meeting_data = {
            'client_id': client_id,
            'client_name': client_name,
            'date': request.form.get('meeting_date'),
            'time': request.form.get('meeting_time'),
            'participants': request.form.get('participants'),
            'topics': request.form.get('topics'),
            'decisions': request.form.get('decisions'),
            'next_steps': request.form.get('next_steps')
        }
        
        # Salva usando o serviço de atas
        if meeting_service:
            meeting_id = meeting_service.save_meeting(meeting_data)
            if meeting_id:
                flash(f'✅ Ata de reunião {meeting_id} registrada com sucesso para {client_name}!', 'success')
            else:
                flash('❌ Erro ao salvar ata de reunião', 'error')
        else:
            # Fallback - salva localmente (simulação)
            flash(f'⚠️ Ata de reunião registrada localmente para {client_name} (funcionalidade limitada)', 'warning')
        
    except Exception as e:
        flash(f'❌ Erro ao salvar ata de reunião: {str(e)}', 'error')
        print(f"❌ Erro detalhado: {e}")
    
    return redirect(url_for('index'))

@app.route('/client/<client_id>/meetings')
@login_required
def view_client_meetings(client_id):
    """Visualizar todas as atas de um cliente"""
    try:
        client = storage_service.get_client(client_id)
        if not client:
            flash('Cliente não encontrado', 'error')
            return redirect(url_for('index'))
        
        meetings = []
        if meeting_service:
            meetings = meeting_service.get_client_meetings(client_id)
        
        return render_template('client_meetings.html', client=client, meetings=meetings)
        
    except Exception as e:
        flash(f'Erro ao carregar atas do cliente: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/meetings')
@login_required
def all_meetings():
    """Visualizar todas as atas de reunião"""
    try:
        meetings = []
        if meeting_service:
            meetings = meeting_service.get_all_meetings()
        
        return render_template('all_meetings.html', meetings=meetings)
        
    except Exception as e:
        flash(f'Erro ao carregar atas: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/debug-render')
def debug_render():
    """Rota de debug para verificar configurações no Render"""
    try:
        import json
        import sys
        
        debug_info = {
            'timestamp': datetime.now().isoformat(),
            'python_version': sys.version,
            'environment': {}
        }
        
        # Verificar variáveis de ambiente
        env_vars = ['GOOGLE_SERVICE_ACCOUNT_JSON', 'GOOGLE_SHEETS_ID', 'FLASK_ENV']
        for var in env_vars:
            value = os.environ.get(var)
            if var == 'GOOGLE_SERVICE_ACCOUNT_JSON' and value:
                # Não expor credenciais completas
                debug_info['environment'][var] = f"Presente ({len(value)} chars)"
                try:
                    creds_info = json.loads(value)
                    debug_info['environment'][f'{var}_parsed'] = {
                        'project_id': creds_info.get('project_id', 'N/A'),
                        'client_email': creds_info.get('client_email', 'N/A'),
                        'keys_available': list(creds_info.keys())
                    }
                except:
                    debug_info['environment'][f'{var}_parsed'] = "Erro ao parsear JSON"
            else:
                debug_info['environment'][var] = value or "Não encontrada"
        
        # Testar importações
        debug_info['imports'] = {}
        try:
            from google.oauth2.service_account import Credentials
            debug_info['imports']['google_oauth2'] = "OK"
        except ImportError as e:
            debug_info['imports']['google_oauth2'] = f"ERRO: {e}"
        
        try:
            from googleapiclient.discovery import build
            debug_info['imports']['googleapiclient'] = "OK"
        except ImportError as e:
            debug_info['imports']['googleapiclient'] = f"ERRO: {e}"
        
        # Testar serviço
        debug_info['service_test'] = {}
        try:
            if USE_GOOGLE_SHEETS and storage_service:
                debug_info['service_test']['storage_type'] = "Google Sheets"
                debug_info['service_test']['service_initialized'] = bool(storage_service.service)
                debug_info['service_test']['spreadsheet_id'] = storage_service.spreadsheet_id
                
                # Testar busca de clientes
                clients = storage_service.get_clients()
                debug_info['service_test']['clients_count'] = len(clients)
                
                if clients:
                    first_client = clients[0]
                    debug_info['service_test']['first_client'] = {
                        'name': first_client.get('nomeEmpresa', 'N/A'),
                        'id': first_client.get('id', 'N/A'),
                        'keys_count': len(first_client.keys())
                    }
            else:
                debug_info['service_test']['storage_type'] = "Local"
                
        except Exception as e:
            debug_info['service_test']['error'] = str(e)
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        })

@app.route('/debug-client/<client_id>')
def debug_client_search(client_id):
    """Rota específica para debug de busca de cliente"""
    try:
        debug_info = {
            'timestamp': datetime.now().isoformat(),
            'client_id_received': client_id,
            'client_id_type': str(type(client_id)),
            'client_id_length': len(str(client_id)),
            'storage_type': 'Google Sheets' if USE_GOOGLE_SHEETS else 'Local'
        }
        
        # Teste 1: Buscar cliente específico
        print(f"🔍 [DEBUG_CLIENT] Buscando cliente: {client_id}")
        client = storage_service.get_client(client_id)
        debug_info['client_found'] = client is not None
        
        if client:
            debug_info['client_data'] = {
                'nome': client.get('nomeEmpresa', 'N/A'),
                'id': client.get('id', 'N/A'),
                'id_type': str(type(client.get('id'))),
                'fields_count': len(client.keys()),
                'has_cnpj': bool(client.get('cnpj')),
                'has_razao_social': bool(client.get('razaoSocialReceita'))
            }
        else:
            debug_info['client_data'] = None
            
            # Teste 2: Listar todos os clientes para comparar IDs
            all_clients = storage_service.get_clients()
            debug_info['total_clients'] = len(all_clients)
            debug_info['sample_client_ids'] = []
            
            for i, c in enumerate(all_clients[:10]):  # Primeiros 10
                debug_info['sample_client_ids'].append({
                    'index': i,
                    'name': c.get('nomeEmpresa', 'N/A'),
                    'id': c.get('id', 'N/A'),
                    'id_matches': str(c.get('id', '')).strip() == str(client_id).strip()
                })
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'timestamp': datetime.now().isoformat(),
            'client_id': client_id
        })

if __name__ == '__main__':
    print("🚀 Iniciando aplicação Flask...")
    print(f"📊 Armazenamento: {'Google Sheets' if USE_GOOGLE_SHEETS else 'Local'}")
    app.run(debug=True, host='0.0.0.0', port=5000)


